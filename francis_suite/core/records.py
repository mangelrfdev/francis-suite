"""
core/records.py

Core data structures for the Francis Suite record system.
Records are structured collections of data with defined schemas.

Created by:   record-create
Populated by: record-add
Inspected by: record-last-added, record-count
Persisted by: record-save, record-save-meta

The record lives in the global context as a shared-box.
All hands access it by name — just like any other shared-box.

FVariable is imported from core/base.py to avoid circular imports
with core/variables.py — both modules need the same FVariable class.

Metadata system:
    Private metadata — always generated automatically, never fails
    Public metadata  — optional, only if <record-metadata> tag is present
"""

from __future__ import annotations
import copy
import hashlib
import html as html_module
import io
import json
import csv
import os
import uuid
import re
import sys
import socket
import platform
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
from datetime import datetime, timezone
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from lxml import etree
from francis_suite.core.base import FVariable


@dataclass
class ExportCustomAttrSpec:
    """Cross-format export key/value from <record-export-attr>."""

    name_raw: str
    value_raw: str
    required: bool = False
    show_attribute: bool = True


@dataclass
class ExportRootAttrSpec:
    """Root-level export (e.g. <Records> attrs in XML, merged into _export elsewhere)."""

    name_raw: str
    value_raw: str
    required: bool = False
    show_attribute: bool = True
    # Legacy record-xml-root-attr: only <Records> in XML, not JSON _export
    xml_only: bool = False


@dataclass
class ExportRowAttrSpec:
    """Per-row attribute (e.g. on <record> in XML) — from field or static."""

    name_raw: str
    value_raw: str
    from_field: str | None
    required: bool = False
    show_attribute: bool = True


@dataclass
class ExportSystemAttrSpec:
    """Built-in values at save time: session_id, total_records, status_process, etc."""

    name_raw: str
    show_attribute: bool = True


# Legacy aliases (tests / older XML)
XmlRootAttrSpec = ExportRootAttrSpec
XmlRecordAttrSpec = ExportRowAttrSpec


def write_text_atomic(path: Path, text: str, *, encoding: str = "utf-8") -> None:
    """Write text via a temp file in the same directory, then replace (best-effort atomic)."""
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_text(text, encoding=encoding)
    tmp.replace(path)


def write_bytes_atomic(path: Path, data: bytes) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_name(path.name + ".tmp")
    tmp.write_bytes(data)
    tmp.replace(path)


def _sanitize_export_string_cell(value: str) -> str:
    """
    Replace CR/LF/tab runs with a single space and collapse repeated spaces.
    Keeps tabular exports (CSV, etc.) to one physical line per row.
    """
    t = re.sub(r"[\r\n\t]+", " ", value)
    return re.sub(r" {2,}", " ", t)


def _sanitize_export_scalar_for_flat(v: Any) -> Any:
    if isinstance(v, str):
        return _sanitize_export_string_cell(v)
    return v


def _sanitize_nested_export_value(obj: Any) -> Any:
    """Deep-apply string cell sanitization for nested record rows (allow-nested)."""
    if isinstance(obj, str):
        return _sanitize_export_string_cell(obj)
    if isinstance(obj, dict):
        return {k: _sanitize_nested_export_value(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_nested_export_value(v) for v in obj]
    return obj


def _strip_shared_group_prefix_from_flat(flat: dict[str, Any]) -> dict[str, Any]:
    """
    When every key is dotted and shares one prefix (e.g. listing.*), drop that segment
    so exports use workflow_key instead of listing.workflow_key.
    If keys mix dotted and non-dotted, or use multiple prefixes, return flat unchanged.
    """
    if not flat:
        return flat
    keys = list(flat.keys())
    if not all("." in k for k in keys):
        return flat
    prefixes = {k.split(".", 1)[0] for k in keys}
    if len(prefixes) != 1:
        return flat
    p = next(iter(prefixes))
    out: dict[str, Any] = {}
    for k, v in flat.items():
        rest = k[len(p) + 1 :] if k.startswith(p + ".") else k
        out[rest] = v
    return out


def _session_completed_for_export(session: Any) -> bool:
    """
    True when export should treat the workflow as successfully finished for show-attribute rules.

    During the last top-level hand, status is still RUNNING but the run will complete right after;
    that case is treated as completed so hidden export fields stay hidden on success.
    """
    if session is None:
        return False
    try:
        st = getattr(session, "status", None)
        val = getattr(st, "value", None) if st is not None else None
    except Exception:
        return False
    if val == "completed":
        return True
    if val == "running" and getattr(session, "_export_final_hand", False):
        return True
    return False


def should_emit_export_show_attribute(show_attribute: bool, session: Any) -> bool:
    """
    show-attribute=false: omit when the export is considered successful/completed; otherwise emit.
    show-attribute=true: always emit.
    """
    if show_attribute:
        return True
    if session is None:
        return True
    return not _session_completed_for_export(session)


# ---------------------------------------------------------------------------
# Francis Suite version
# ---------------------------------------------------------------------------

FRANCIS_SUITE_VERSION = "0.1.0"


# ---------------------------------------------------------------------------
# XML export helpers (record-save format xml)
# ---------------------------------------------------------------------------

def _sanitize_xml_tag(name: str) -> str:
    s = re.sub(r"[^0-9A-Za-z_\-.]", "_", str(name))
    if s and s[0].isdigit():
        s = "_" + s
    return s or "field"


def _append_xml_from_value(parent: etree._Element, key: str, value: Any) -> None:
    tag = _sanitize_xml_tag(key)
    if isinstance(value, dict):
        el = etree.SubElement(parent, tag)
        for k, v in value.items():
            _append_xml_from_value(el, k, v)
    elif isinstance(value, (list, tuple)):
        el = etree.SubElement(parent, tag)
        for i, item in enumerate(value):
            if isinstance(item, dict):
                sub = etree.SubElement(el, "item", index=str(i))
                for k, v in item.items():
                    _append_xml_from_value(sub, k, v)
            else:
                sub = etree.SubElement(el, "item", index=str(i))
                if item is None:
                    sub.text = ""
                elif isinstance(item, str):
                    sub.text = _sanitize_export_string_cell(item)
                else:
                    sub.text = str(item)
    else:
        el = etree.SubElement(parent, tag)
        if value is None:
            el.text = ""
        elif isinstance(value, str):
            el.text = _sanitize_export_string_cell(value)
        else:
            el.text = str(value)


# ---------------------------------------------------------------------------
# Available auto fields for <metadata-field name="..."/>
# These are computed from the record itself at save time
# ---------------------------------------------------------------------------

AUTO_METADATA_FIELDS = {
    # data quality
    "total_rows",
    "rows_completados",
    "rows_con_campos_vacios",
    "rows_fallidos",
    "campos_nulos_total",
    "porcentaje_completitud",
    "filas_duplicadas_por_clave",
    "filas_rechazadas_validacion",
    # performance
    "duracion_segundos",
    "rows_por_segundo",
    # traceability
    "session_id",
    "workflow_path",
    "francis_suite_version",
    "hostname",
    "sistema_operativo",
    "python_version",
    "status",
    "error",
    "inicio",
    "fin",
}


# ---------------------------------------------------------------------------
# Valid field types
# ---------------------------------------------------------------------------

VALID_TYPES = {
    "string",
    "integer",
    "decimal",
    "boolean",
    "date",
    "datetime",
    "url",
    "email",
    "uuid",
}

_TRUE_VALUES  = {"true", "yes", "si", "sí", "1", "verdadero"}
_FALSE_VALUES = {"false", "no", "0", "falso"}


# ---------------------------------------------------------------------------
# FRecordField
# ---------------------------------------------------------------------------

class FRecordField:
    """Defines a single field in a record schema."""

    def __init__(
        self,
        name: str,
        field_type: str,
        required: bool = True,
        null_if_empty: bool = False,
        precision: str | None = None,
    ) -> None:
        if field_type not in VALID_TYPES:
            raise ValueError(
                f"[RECORD] invalid field type '{field_type}' for field '{name}'. "
                f"Valid types: {', '.join(sorted(VALID_TYPES))}"
            )

        self.name          = name
        self.type          = field_type
        self.required      = required
        self.null_if_empty = null_if_empty
        self.precision     = precision

    def normalize(self, raw_value: Any) -> Any:
        """Normalize and validate a raw value according to this field's type."""
        value = str(raw_value).strip() if raw_value is not None else ""

        # Workflow literal "NULL" in <box-def> → explicit None for optional fields (JSON null).
        if value.upper() == "NULL":
            if not self.required:
                return None
            raise ValueError(
                f"[RECORD] field '{self.name}' is required — literal NULL is not allowed"
            )

        if not value:
            if self.type == "uuid":
                return str(uuid.uuid4())
            if self.required:
                raise ValueError(
                    f"[RECORD] field '{self.name}' is required but got empty value"
                )
            return None if self.null_if_empty else ""

        if self.type == "string":
            return value
        elif self.type == "integer":
            return self._normalize_integer(value)
        elif self.type == "decimal":
            return self._normalize_decimal(value)
        elif self.type == "boolean":
            return self._normalize_boolean(value)
        elif self.type == "date":
            return self._normalize_date(value)
        elif self.type == "datetime":
            return self._normalize_datetime(value)
        elif self.type == "url":
            return self._normalize_url(value)
        elif self.type == "email":
            return self._normalize_email(value)
        elif self.type == "uuid":
            return self._normalize_uuid(value)

        return value

    def _normalize_integer(self, value: str) -> int:
        cleaned = re.sub(r"[^\d\-]", "", value.replace(",", ""))
        if not cleaned or cleaned == "-":
            raise ValueError(
                f"[RECORD] field '{self.name}' expects integer but got '{value}'"
            )
        try:
            return int(cleaned)
        except ValueError:
            raise ValueError(
                f"[RECORD] field '{self.name}' expects integer but got '{value}'"
            )

    def _normalize_decimal(self, value: str) -> str:
        if re.search(r"\d\.\d{3},", value):
            cleaned = re.sub(r"[^\d,\-]", "", value)
            cleaned = cleaned.replace(".", "").replace(",", ".")
        elif re.search(r"\d,\d{1,2}$", value):
            cleaned = re.sub(r"[^\d,\-]", "", value)
            cleaned = cleaned.replace(",", ".")
        else:
            cleaned = re.sub(r"[^\d\.\-]", "", value)

        if not cleaned or cleaned == "-":
            raise ValueError(
                f"[RECORD] field '{self.name}' expects decimal but got '{value}'"
            )

        try:
            decimal_value = Decimal(cleaned)
        except InvalidOperation:
            raise ValueError(
                f"[RECORD] field '{self.name}' expects decimal but got '{value}'"
            )

        if self.precision and self.precision != "raw":
            try:
                places = int(self.precision)
                quantize_str = "0." + "0" * places if places > 0 else "0"
                decimal_value = decimal_value.quantize(
                    Decimal(quantize_str),
                    rounding=ROUND_HALF_UP
                )
            except (ValueError, InvalidOperation):
                raise ValueError(
                    f"[RECORD] invalid precision '{self.precision}' for field '{self.name}'"
                )

        return str(decimal_value)

    def _normalize_boolean(self, value: str) -> bool:
        lower = value.lower().strip()
        if lower in _TRUE_VALUES:
            return True
        if lower in _FALSE_VALUES:
            return False
        raise ValueError(
            f"[RECORD] field '{self.name}' expects boolean but got '{value}'. "
            f"Valid values: {', '.join(sorted(_TRUE_VALUES | _FALSE_VALUES))}"
        )

    def _normalize_date(self, value: str) -> str:
        formats = [
            "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y",
            "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        raise ValueError(
            f"[RECORD] field '{self.name}' expects date but got '{value}'. "
            f"Supported formats: DD/MM/YYYY, YYYY-MM-DD, DD-MM-YYYY"
        )

    def _normalize_datetime(self, value: str) -> str:
        formats = [
            "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
            "%d-%m-%Y %H:%M:%S", "%d-%m-%Y %H:%M",
        ]
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt).strftime("%Y-%m-%dT%H:%M:%S")
            except ValueError:
                continue
        raise ValueError(
            f"[RECORD] field '{self.name}' expects datetime but got '{value}'. "
            f"Supported formats: YYYY-MM-DDTHH:MM:SS, DD/MM/YYYY HH:MM:SS"
        )

    def _normalize_url(self, value: str) -> str:
        if not re.match(r"^https?://", value):
            raise ValueError(
                f"[RECORD] field '{self.name}' expects URL but got '{value}'. "
                f"URL must start with http:// or https://"
            )
        return value

    def _normalize_email(self, value: str) -> str:
        if not re.match(r"^[^@]+@[^@]+\.[^@]+$", value):
            raise ValueError(
                f"[RECORD] field '{self.name}' expects email but got '{value}'"
            )
        return value.lower()

    def _normalize_uuid(self, value: str) -> str:
        try:
            return str(uuid.UUID(value))
        except ValueError:
            raise ValueError(
                f"[RECORD] field '{self.name}' expects UUID but got '{value}'"
            )


# ---------------------------------------------------------------------------
# FRecordGroup
# ---------------------------------------------------------------------------

class FRecordGroup:
    """A named group of fields within a record schema."""

    def __init__(self, name: str, required: bool = True) -> None:
        self.name     = name
        self.required = required
        self.fields: dict[str, FRecordField] = {}

    def add_field(self, field: FRecordField) -> None:
        self.fields[field.name] = field

    def normalize_row(self, raw_row: dict) -> dict:
        result = {}
        for field_name, field in self.fields.items():
            raw_value = raw_row.get(field_name, "")
            result[field_name] = field.normalize(raw_value)
        return result


# ---------------------------------------------------------------------------
# FRecordSchema
# ---------------------------------------------------------------------------

class FRecordSchema:
    """Complete schema for a record collection."""

    def __init__(self, name: str) -> None:
        self.name   = name
        self.groups: dict[str, FRecordGroup] = {}

        # public metadata — optional, set by <record-metadata>
        # list of {"name": str, "value": str | None}
        # value=None means auto-computed at save time
        self.public_metadata_fields: list[dict] = []
        self.has_public_metadata: bool = False

        # record-key — optional, from <record-key><key-field name="..."/></record-key>
        # Stored as "group.field" keys in declaration order
        self.record_key_keys: list[str] = []

        # Export augmentation — from <record-create>; resolved at record-save
        self.export_custom_specs: list[ExportCustomAttrSpec] = []
        self.export_root_specs: list[ExportRootAttrSpec] = []
        self.export_row_specs: list[ExportRowAttrSpec] = []
        self.export_system_specs: list[ExportSystemAttrSpec] = []

        # Optional NDJSON journal — one appended line per successful add_row (crash-safe incremental)
        self.journal_path: str | None = None
        self.journal_fsync: bool = False

        # strict (default): invalid row raises. collect-errors: skip row, record in validation_errors
        self.record_validation_mode: str = "strict"

    def add_group(self, group: FRecordGroup) -> None:
        self.groups[group.name] = group

    def add_public_metadata_field(self, name: str, value: str | None = None) -> None:
        """Add a field to public metadata. value=None means auto-computed."""
        self.has_public_metadata = True
        self.public_metadata_fields.append({"name": name, "value": value})

    def normalize_row(self, raw_row: dict) -> dict:
        result = {}
        for group_name, group in self.groups.items():
            group_raw = {}
            for field_name in group.fields:
                key = f"{group_name}.{field_name}"
                group_raw[field_name] = raw_row.get(key, "")
            result[group_name] = group.normalize_row(group_raw)
        return result

    def add_record_key_field(self, field_spec: str) -> None:
        """
        Register one field that participates in the duplicate-detection key.
        field_spec: bare field name (must be unique across groups) or "group.field".
        """
        resolved = self._resolve_key_field(field_spec.strip())
        if resolved in self.record_key_keys:
            raise ValueError(
                f"[RECORD] duplicate key-field '{field_spec}' in <record-key>"
            )
        self.record_key_keys.append(resolved)

    def _resolve_key_field(self, spec: str) -> str:
        if not self.groups:
            raise ValueError(
                "[RECORD] <record-key> must be declared after <record-set-group> fields exist"
            )
        if "." in spec:
            group_name, field_name = spec.split(".", 1)
            group_name = group_name.strip()
            field_name = field_name.strip()
            if group_name not in self.groups:
                raise ValueError(
                    f"[RECORD] key-field references unknown group '{group_name}'"
                )
            if field_name not in self.groups[group_name].fields:
                raise ValueError(
                    f"[RECORD] key-field references unknown field '{field_name}' "
                    f"in group '{group_name}'"
                )
            return f"{group_name}.{field_name}"

        matches: list[str] = []
        for gname, group in self.groups.items():
            if spec in group.fields:
                matches.append(f"{gname}.{spec}")
        if len(matches) == 1:
            return matches[0]
        if len(matches) == 0:
            raise ValueError(
                f"[RECORD] key-field '{spec}' not found in any record-set-group"
            )
        raise ValueError(
            f"[RECORD] key-field '{spec}' is ambiguous — use 'group.field' "
            f"(matches: {', '.join(matches)})"
        )

    def resolve_flat_field_path(self, bare_or_qualified: str) -> str:
        """Resolve a field reference to flattened key used by _flatten (e.g. book.record_key)."""
        s = bare_or_qualified.strip()
        if not s:
            raise ValueError("[RECORD] from-field is empty")
        if "." in s:
            g, f = s.split(".", 1)
            g, f = g.strip(), f.strip()
            if g not in self.groups or f not in self.groups[g].fields:
                raise ValueError(
                    f"[RECORD] from-field '{bare_or_qualified}' does not match a declared field"
                )
            return f"{g}.{f}"
        matches: list[str] = []
        for gname, group in self.groups.items():
            if s in group.fields:
                matches.append(f"{gname}.{s}")
        if len(matches) == 1:
            return matches[0]
        if len(matches) == 0:
            raise ValueError(
                f"[RECORD] from-field '{bare_or_qualified}' not found — use group.field"
            )
        raise ValueError(
            f"[RECORD] from-field '{bare_or_qualified}' is ambiguous — use group.field "
            f"(matches: {', '.join(matches)})"
        )


# ---------------------------------------------------------------------------
# FRecord
# ---------------------------------------------------------------------------

class FRecord(FVariable):
    """
    A structured collection of records with a defined schema.
    Inherits FVariable from core/base.py.

    Metadata system:
        Private metadata — always generated automatically, never fails
        Public metadata  — optional, only if schema.has_public_metadata is True
    """

    def __init__(self, schema: FRecordSchema) -> None:
        self._schema        = schema
        self._rows:         list[dict] = []
        self._duplicate_rows: list[dict] = []
        self._validation_error_rows: list[dict[str, Any]] = []
        self._rows_failed:  int = 0
        self._private_meta: dict = {}
        self._seen_record_key_hashes: set[str] = set()
        self._journal_finalized: bool = False
        # Set by <record-save-metadata path="..."/> — written in runtime after session completes
        # so status/duration/fin reflect the final session state.
        self._deferred_private_metadata_path: str | None = None

        # RAM tracking with psutil if available
        self._ram_samples:  list[float] = []
        self._ram_peak:     float | None = None
        self._sample_ram()

    def _sample_ram(self) -> None:
        """Sample current RAM usage. Silently skips if psutil not available."""
        try:
            import psutil
            import os
            process   = psutil.Process(os.getpid())
            ram_mb    = process.memory_info().rss / 1024 / 1024
            self._ram_samples.append(ram_mb)
            if self._ram_peak is None or ram_mb > self._ram_peak:
                self._ram_peak = ram_mb
        except Exception:
            pass

    @property
    def name(self) -> str:
        return self._schema.name

    @property
    def count(self) -> int:
        return len(self._rows)

    def register_deferred_private_metadata_path(self, path: str) -> None:
        """
        Custom path for private metadata JSON. The runtime writes this file once after the
        session ends (success or failure) so fields like status, fin, duracion_segundos match
        the final session instead of mid-workflow snapshots.
        """
        p = (path or "").strip()
        self._deferred_private_metadata_path = p if p else None

    @property
    def deferred_private_metadata_path(self) -> str | None:
        return self._deferred_private_metadata_path

    @property
    def duplicate_count(self) -> int:
        """Rows skipped because of duplicate <record-key> (subsequent occurrences)."""
        return len(self._duplicate_rows)

    @property
    def duplicate_rows(self) -> list[dict]:
        """Copy of rows that were skipped as duplicate keys (for record-save-duplicates)."""
        return list(self._duplicate_rows)

    @property
    def has_record_key(self) -> bool:
        """True if <record-key> was declared (duplicate detection enabled)."""
        return bool(self._schema.record_key_keys)

    @property
    def validation_error_count(self) -> int:
        """Rows rejected by schema validation when record-validation is collect-errors."""
        return len(self._validation_error_rows)

    @property
    def validation_error_rows(self) -> list[dict[str, Any]]:
        """Copy of validation error entries (error message + raw_row from record-add)."""
        return list(self._validation_error_rows)

    @property
    def last_row(self) -> dict | None:
        return self._rows[-1] if self._rows else None

    def validation_error_export_rows(self) -> list[dict[str, Any]]:
        """
        Flat dicts for record-save-validation-errors: validation_error + raw_* field keys.
        """
        out: list[dict[str, Any]] = []
        for entry in self._validation_error_rows:
            row: dict[str, Any] = {"validation_error": entry["error"]}
            raw = entry.get("raw_row") or {}
            for k, v in raw.items():
                row[f"raw_{k}"] = "" if v is None else str(v)
            out.append(row)
        return out

    def add_row(self, raw_row: dict, session: Any = None) -> dict | None:
        """
        Normalize and add a row to the collection.
        Returns None if <record-key> is configured and this row duplicates an existing key.
        Returns None if record-validation is collect-errors and normalization raises (row stored as validation error).
        session: optional — used for journal lines (session id, status, workflow name).
        """
        try:
            normalized = self._schema.normalize_row(raw_row)
        except ValueError as e:
            if self._schema.record_validation_mode == "collect-errors":
                msg = str(e)
                self._validation_error_rows.append({"error": msg, "raw_row": dict(raw_row)})
                short = msg if len(msg) <= 120 else msg[:117] + "..."
                print(f"[RECORD] validation error — skipping row ({short})")
                return None
            raise
        if self._schema.record_key_keys:
            key_hash = self._make_record_key_hash(normalized)
            if key_hash in self._seen_record_key_hashes:
                short = key_hash[:16]
                print(f"[RECORD] duplicate key — skipping (key: {short})")
                self._duplicate_rows.append(normalized)
                return None
            self._seen_record_key_hashes.add(key_hash)
        self._rows.append(normalized)
        self._sample_ram()
        self._append_journal_line(normalized, session=session)
        return normalized

    def write_journal_header_if_needed(self, session: Any) -> None:
        """
        Write a single journal_header line when the journal file is new/empty.
        Call once from record-create after the FRecord is registered.
        """
        raw_path = self._schema.journal_path
        if not raw_path:
            return
        path = Path(raw_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        if path.exists() and path.stat().st_size > 0:
            return
        header = {
            "_type": "journal_header",
            "francis_suite_version": FRANCIS_SUITE_VERSION,
            "record_name": self.name,
            "session_id": getattr(session, "id", None) if session else None,
            "workflow_name": getattr(session, "workflow_name", None) if session else None,
            "status": "running",
            "started_at": datetime.now(timezone.utc).isoformat(),
        }
        with open(path, "w", encoding="utf-8") as f:
            f.write(json.dumps(header, ensure_ascii=False, default=str) + "\n")
            if self._schema.journal_fsync:
                f.flush()
                os.fsync(f.fileno())

    def finalize_journal(self, session: Any) -> None:
        """
        Append a process summary line when the workflow ends (success or failure).
        Safe to call multiple times — only runs if journal_path is set.
        """
        if self._journal_finalized:
            return
        raw_path = self._schema.journal_path
        if not raw_path:
            return
        path = Path(raw_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        status_val = "unknown"
        err_msg: str | None = None
        if session is not None:
            try:
                st = session.status
                status_val = getattr(st, "value", None) or str(st)
            except Exception:
                status_val = "unknown"
            try:
                if session.error:
                    err_msg = str(session.error)
            except Exception:
                pass
        summary = {
            "_type": "process",
            "status": status_val,
            "session_id": getattr(session, "id", None) if session else None,
            "workflow_name": getattr(session, "workflow_name", None) if session else None,
            "record_name": self.name,
            "rows_committed": len(self._rows),
            "francis_suite_version": FRANCIS_SUITE_VERSION,
            "error": err_msg,
            "finished_at": datetime.now(timezone.utc).isoformat(),
        }
        with open(path, "a", encoding="utf-8") as f:
            f.write(json.dumps(summary, ensure_ascii=False, default=str) + "\n")
            if self._schema.journal_fsync:
                f.flush()
                os.fsync(f.fileno())
        self._journal_finalized = True

    def _append_journal_line(self, normalized: dict, session: Any = None) -> None:
        """Append one NDJSON line if journal_path was set on record-create."""
        raw_path = self._schema.journal_path
        if not raw_path:
            return
        path = Path(raw_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        flat = self._flatten(normalized)
        payload: dict[str, Any] = {
            "_type": "record",
            "row_index": len(self._rows),
            "francis_suite_version": FRANCIS_SUITE_VERSION,
            "data": flat,
        }
        if session is not None:
            payload["session_id"] = getattr(session, "id", None)
            try:
                st = session.status
                payload["status"] = getattr(st, "value", None)
            except Exception:
                payload["status"] = None
            payload["workflow_name"] = getattr(session, "workflow_name", None)
        line = json.dumps(payload, ensure_ascii=False, default=str) + "\n"
        with open(path, "a", encoding="utf-8") as f:
            f.write(line)
            if self._schema.journal_fsync:
                f.flush()
                os.fsync(f.fileno())

    def _sanitize_export_for_format(self, fmt: str, xa: dict[str, str]) -> dict[str, str]:
        """
        Drop or fix export augmentation keys so each format only gets compatible values.
        Row payload is unchanged — this only affects _export / root augmentation dicts.
        """
        if not xa:
            return xa
        out: dict[str, str] = {}
        for k, v in xa.items():
            sv = "" if v is None else str(v)
            # No raw newlines in CSV/TSV/HTML comments, XML attributes, parquet metadata JSON, etc.
            if fmt in ("csv", "txt", "html", "excel", "xml", "json", "ndjson", "parquet"):
                sv = sv.replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()
            out[k] = sv
        if fmt == "parquet":
            blob = json.dumps(out, ensure_ascii=False, default=str)
            if len(blob.encode("utf-8")) > 65536:
                return {}
        return out

    def _make_record_key_hash(self, normalized: dict) -> str:
        """Stable SHA-256 hash of key field values in schema order."""
        parts: list[tuple[str, Any]] = []
        for gk in self._schema.record_key_keys:
            g_name, f_name = gk.split(".", 1)
            group_data = normalized.get(g_name, {})
            val = group_data.get(f_name) if isinstance(group_data, dict) else None
            parts.append((gk, val))
        payload = json.dumps(parts, sort_keys=True, default=str, ensure_ascii=False)
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def add_private_metadata(self, name: str, value: str) -> None:
        """Add a custom field to private metadata."""
        self._private_meta[name] = value

    # --- Data quality stats ---

    def _compute_quality(self) -> dict:
        """Compute data quality stats from rows."""
        if not self._rows:
            return {
                "total_rows":              0,
                "rows_completados":        0,
                "rows_con_campos_vacios":  0,
                "rows_fallidos":           self._rows_failed,
                "campos_nulos_total":      0,
                "porcentaje_completitud":  None,
            }

        rows_con_vacios = 0
        campos_nulos    = 0

        for row in self._rows:
            flat = self._flatten(row)
            tiene_vacio = False
            for v in flat.values():
                if v is None or v == "":
                    campos_nulos += 1
                    tiene_vacio   = True
            if tiene_vacio:
                rows_con_vacios += 1

        total             = len(self._rows)
        rows_completados  = total - rows_con_vacios
        completitud       = round((rows_completados / total) * 100, 2) if total > 0 else None

        return {
            "total_rows":              total,
            "rows_completados":        rows_completados,
            "rows_con_campos_vacios":  rows_con_vacios,
            "rows_fallidos":           self._rows_failed,
            "campos_nulos_total":      campos_nulos,
            "porcentaje_completitud":  completitud,
        }

    # --- Private metadata generation ---

    def build_private_metadata(self, session=None) -> dict:
        """
        Build complete private metadata.
        Always returns all fields — missing values are None.
        Never raises — designed to work even if session is None or failed.
        """
        self._sample_ram()

        quality = self._compute_quality()

        # performance
        duracion        = None
        rows_por_segundo = None
        inicio          = None
        fin             = None
        session_id      = None
        workflow_path   = None
        status          = None
        error           = None

        if session is not None:
            try:
                session_id    = session.id
                workflow_path = session.workflow_name
                status        = session.status.value
                error         = str(session.error) if session.error else None
                inicio        = session.started_at.isoformat() if session.started_at else None
                fin           = session.finished_at.isoformat() if session.finished_at else None
                duracion      = session.duration
                if duracion and quality["total_rows"]:
                    rows_por_segundo = round(quality["total_rows"] / duracion, 2)
            except Exception:
                pass

        # RAM
        ram_peak    = self._ram_peak
        ram_promedio = round(sum(self._ram_samples) / len(self._ram_samples), 2) if self._ram_samples else None

        # system info
        try:
            hostname = socket.gethostname()
        except Exception:
            hostname = None

        try:
            so = platform.system() + "-" + platform.release()
        except Exception:
            so = None

        try:
            py_version = sys.version.split()[0]
        except Exception:
            py_version = None

        meta = {
            # traceability
            "session_id":              session_id,
            "workflow_path":           workflow_path,
            "francis_suite_version":   FRANCIS_SUITE_VERSION,
            "hostname":                hostname,
            "sistema_operativo":       so,
            "python_version":          py_version,
            "status":                  status,
            "error":                   error,
            "inicio":                  inicio,
            "fin":                     fin,

            # performance
            "duracion_segundos":               duracion,
            "ram_peak_mb":                     ram_peak,
            "ram_promedio_mb":                 ram_promedio,
            "rows_por_segundo":                rows_por_segundo,
            "requests_http_total":             None,  # future: httpx tracker
            "requests_http_fallidas":          None,  # future: httpx tracker
            "tiempo_promedio_por_request_ms":  None,  # future: httpx tracker

            # data quality
            "total_rows":              quality["total_rows"],
            "rows_completados":        quality["rows_completados"],
            "rows_con_campos_vacios":  quality["rows_con_campos_vacios"],
            "rows_fallidos":           quality["rows_fallidos"],
            "campos_nulos_total":      quality["campos_nulos_total"],
            "porcentaje_completitud":  quality["porcentaje_completitud"],
            "filas_duplicadas_por_clave":     self.duplicate_count,
            "filas_rechazadas_validacion":    self.validation_error_count,

            # scraping specific — populated by <record-private-metadata>
            "paginas_procesadas":      self._private_meta.get("paginas_procesadas", None),
            "paginas_fallidas":        self._private_meta.get("paginas_fallidas", None),
            "urls_visitadas":          self._private_meta.get("urls_visitadas", None),
            "proxies_usados":          None,  # future: proxy system
            "captchas_encontrados":    None,  # future: playwright
            "rate_limits_alcanzados":  None,  # future: httpx tracker
        }

        # add any extra private fields the user added
        for key, value in self._private_meta.items():
            if key not in meta:
                meta[key] = value

        return meta

    # --- Public metadata generation ---

    def build_public_metadata(self, session=None) -> dict | None:
        """
        Build public metadata from declared fields.
        Returns None if no public metadata was declared.
        Only called when status=completed.
        """
        if not self._schema.has_public_metadata:
            return None

        private = self.build_private_metadata(session)
        quality = self._compute_quality()
        all_auto = {**private, **quality}

        result = {}
        for field in self._schema.public_metadata_fields:
            name  = field["name"]
            value = field["value"]

            if value is not None:
                # user provided a fixed value
                result[name] = value
            elif name in all_auto:
                # auto-computed field
                result[name] = all_auto[name]
            else:
                result[name] = None

        return result

    # --- FVariable interface ---

    def to_string(self) -> str:
        return f"[RECORD:{self.name}:{self.count}]"

    def to_display(self) -> str:
        return self.to_string()

    def to_list(self) -> list:
        return list(self._rows)

    def is_empty(self) -> bool:
        return len(self._rows) == 0

    # --- Persistence ---

    def save(
        self,
        format: str,
        path: str,
        *,
        data_rows: list[dict] | None = None,
        include_metadata: bool = False,
        clean_data: bool = False,
        allow_nested: bool = False,
        allow_prefix: bool = False,
        session=None,
        sheet_name: str = "Data",
        metadata_sheet_name: str = "Metadata",
        html_title: str | None = None,
        export_augmentation: dict[str, str] | None = None,
        xml_include_root_workflow: bool = True,
        xml_include_root_total_records: bool = True,
        xml_include_root_session_id: bool = False,
        xml_include_root_francis_version: bool = False,
        xml_include_root_exported_at: bool = False,
        xml_include_record_workflow: bool = True,
        xml_include_record_key: bool = True,
        xml_root_extra_attrs: dict[str, str] | None = None,
        xml_record_extra_attrs: dict[str, str] | None = None,
        xml_record_attr_specs_resolved: list[dict] | None = None,
    ) -> None:
        """
        Persist the record collection to disk.

        Args:
            format:                 json, csv, ndjson, xml, html, txt, excel, xlsx, parquet
            path:                   output file path
            data_rows:              if set, serialize these rows instead of primary _rows (duplicate-key
                                    export via record-save-duplicates). include_metadata must be False.
            include_metadata:       if True, embeds public metadata where the format supports it
            clean_data:             if True, omit export/session lines and public metadata from the
                                    serialized file (CSV # lines, NDJSON export/metadata lines,
                                    JSON _export/_metadata wrappers, HTML/Excel export sections,
                                    Parquet francis_export key, etc.). Overrides include-metadata
                                    for output framing.
            allow_nested:           if True, json/ndjson rows keep nested group objects; tabular formats
                                    use dotted keys (listing.field) like allow-prefix. Default: false.
            allow_prefix:           if True, flat keys keep the group prefix (listing.field). Ignored for
                                    json/ndjson when allow_nested is True. Default: false.
            session:                FrancisSession — workflow name and metadata fields
            sheet_name:             excel — main data sheet name
            metadata_sheet_name:    excel — second sheet for public metadata (if include_metadata)
            html_title:             html — <title> and main heading (default: workflow name)
            export_augmentation:    optional key/value (from record-create or caller) — applied per format
            xml_record_attr_specs_resolved: per-row <record> attr specs, pre-resolved
            xml_*:                  only for format xml — see _save_xml
        """
        if data_rows is not None and include_metadata:
            raise ValueError(
                "[RECORD] include-metadata is not supported when saving duplicate-key rows "
                "(use <record-save> for public metadata; <record-save-duplicates> omits it)."
            )

        clean_data_flag = clean_data
        effective_include_metadata = include_metadata and not clean_data_flag

        tabular_use_prefix = bool(allow_nested or allow_prefix)
        json_ndjson_shape = (
            "nested"
            if allow_nested
            else ("prefix" if allow_prefix else "clean")
        )

        effective_rows = self._rows if data_rows is None else data_rows
        if not effective_rows:
            return

        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        fmt = format.lower().strip()
        xa = dict(export_augmentation or {})
        if fmt == "xml":
            if xml_include_root_session_id and "session_id" not in xa:
                sid = getattr(session, "id", None) if session else None
                xa["session_id"] = "" if sid is None else str(sid)
            if xml_include_root_francis_version and "francis_suite_version" not in xa:
                xa["francis_suite_version"] = FRANCIS_SUITE_VERSION
            if xml_include_root_exported_at and "exported_at" not in xa:
                xa["exported_at"] = datetime.now(timezone.utc).isoformat()

        xa = self._sanitize_export_for_format(fmt, xa)
        if clean_data_flag:
            xa = {}

        include_export_wrapper = (export_augmentation is not None) and not clean_data_flag

        if data_rows is None:
            row_suffix = ""
        elif (
            effective_rows
            and isinstance(effective_rows[0], dict)
            and "validation_error" in effective_rows[0]
        ):
            row_suffix = " (validation-error rows)"
        else:
            row_suffix = " (duplicate-key rows)"

        if fmt == "json":
            self._save_json(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                session=session,
                export_augmentation_payload=xa,
                include_export_wrapper=include_export_wrapper,
                export_row_shape=json_ndjson_shape,
            )
        elif fmt == "csv":
            self._save_csv(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                export_augmentation=xa,
                tabular_use_prefix=tabular_use_prefix,
            )
        elif fmt == "ndjson":
            self._save_ndjson(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                session=session,
                export_augmentation_payload=xa,
                include_export_wrapper=include_export_wrapper,
                export_row_shape=json_ndjson_shape,
            )
        elif fmt == "xml":
            self._save_xml(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                session=session,
                include_root_workflow=xml_include_root_workflow,
                include_root_total_records=xml_include_root_total_records,
                include_record_workflow=xml_include_record_workflow,
                include_record_key=xml_include_record_key,
                root_extra_attrs=xml_root_extra_attrs or {},
                record_extra_attrs=xml_record_extra_attrs or {},
                export_augmentation=xa,
                xml_record_attr_specs_resolved=xml_record_attr_specs_resolved or [],
            )
        elif fmt == "html":
            self._save_html(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                session=session,
                html_title=html_title,
                export_augmentation=xa,
                tabular_use_prefix=tabular_use_prefix,
            )
        elif fmt == "txt":
            self._save_txt(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                export_augmentation=xa,
                tabular_use_prefix=tabular_use_prefix,
            )
        elif fmt in ("excel", "xlsx"):
            self._save_excel(
                output_path,
                effective_rows,
                include_metadata=effective_include_metadata,
                session=session,
                sheet_name=sheet_name,
                metadata_sheet_name=metadata_sheet_name,
                export_augmentation=xa,
                tabular_use_prefix=tabular_use_prefix,
            )
        elif fmt == "parquet":
            self._save_parquet(
                output_path,
                effective_rows,
                export_augmentation=xa,
                tabular_use_prefix=tabular_use_prefix,
            )
        else:
            raise ValueError(
                f"[RECORD] unsupported format '{format}'. "
                f"Valid formats: json, csv, ndjson, xml, html, txt, excel, xlsx, parquet"
            )

        print(
            f"[RECORD] saved {len(effective_rows)} rows to '{output_path}' as {fmt}{row_suffix}"
        )

    def save_meta(self, path: str, session=None) -> None:
        """
        Persist only the private metadata to disk as JSON.
        Always works — even if the session failed.
        """
        output_path = Path(path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        meta = self.build_private_metadata(session)

        text = json.dumps(meta, ensure_ascii=False, indent=2, default=str)
        write_text_atomic(output_path, text)

        print(f"[RECORD] saved metadata to '{output_path}'")

    def _save_json(
        self,
        path: Path,
        rows: list[dict],
        include_metadata: bool = False,
        session=None,
        *,
        export_augmentation_payload: dict[str, str] | None = None,
        include_export_wrapper: bool = False,
        export_row_shape: str = "clean",
    ) -> None:
        xa = dict(export_augmentation_payload or {})
        export_rows = self._rows_for_json_ndjson_export(rows, shape=export_row_shape)
        if include_metadata:
            public_meta = self.build_public_metadata(session)
            output: dict | list = {
                "_metadata": public_meta or {},
                "data": export_rows,
            }
            if include_export_wrapper:
                output["_export"] = xa
            elif xa:
                output["_export"] = xa
        elif include_export_wrapper:
            output = {"_export": xa, "data": export_rows}
        else:
            output = export_rows

        text = json.dumps(output, ensure_ascii=False, indent=2, default=str)
        write_text_atomic(path, text)

    def _save_ndjson(
        self,
        path: Path,
        rows: list[dict],
        include_metadata: bool = False,
        session=None,
        *,
        export_augmentation_payload: dict[str, str] | None = None,
        include_export_wrapper: bool = False,
        export_row_shape: str = "clean",
    ) -> None:
        xa = dict(export_augmentation_payload or {})
        lines: list[str] = []
        if include_export_wrapper:
            lines.append(json.dumps({"_type": "export", **xa}, ensure_ascii=False, default=str))
        elif xa:
            lines.append(json.dumps({"_type": "export", **xa}, ensure_ascii=False, default=str))
        if include_metadata:
            public_meta = self.build_public_metadata(session)
            if public_meta:
                lines.append(json.dumps({"_type": "metadata", **public_meta}, ensure_ascii=False, default=str))
        shaped = self._rows_for_json_ndjson_export(rows, shape=export_row_shape)
        for row in shaped:
            lines.append(json.dumps(row, ensure_ascii=False, default=str))
        if lines:
            write_text_atomic(path, "\n".join(lines) + "\n")

    def _save_csv(
        self,
        path: Path,
        rows: list[dict],
        include_metadata: bool = False,
        export_augmentation: dict[str, str] | None = None,
        *,
        tabular_use_prefix: bool = False,
    ) -> None:
        if not rows:
            return

        flat_rows_raw = self._rows_for_tabular_export(
            rows, tabular_use_prefix=tabular_use_prefix
        )
        # CSV has no JSON null — empty string for missing / None
        flat_rows = [
            {k: ("" if v is None else v) for k, v in r.items()}
            for r in flat_rows_raw
        ]

        keys: list[str] = []
        for row in flat_rows:
            for key in row:
                if key not in keys:
                    keys.append(key)

        xa = export_augmentation or {}
        buf = io.StringIO()
        if include_metadata and self._schema.has_public_metadata:
            for field in self._schema.public_metadata_fields:
                value = field["value"] or ""
                buf.write(f"# {field['name']}: {value}\n")
        for ek, ev in xa.items():
            buf.write(f"# {ek}: {ev}\n")

        writer = csv.DictWriter(
            buf, fieldnames=keys, extrasaction="ignore", lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows(flat_rows)
        write_text_atomic(path, buf.getvalue())

    def _workflow_name(self, session) -> str:
        if session is None:
            return ""
        return str(getattr(session, "workflow_name", "") or "")

    def _save_xml(
        self,
        path: Path,
        rows: list[dict],
        *,
        include_metadata: bool = False,
        session=None,
        include_root_workflow: bool = True,
        include_root_total_records: bool = True,
        include_record_workflow: bool = True,
        include_record_key: bool = True,
        root_extra_attrs: dict[str, str] | None = None,
        record_extra_attrs: dict[str, str] | None = None,
        export_augmentation: dict[str, str] | None = None,
        xml_record_attr_specs_resolved: list[dict] | None = None,
    ) -> None:
        wf = self._workflow_name(session)
        root_attrs: dict[str, str] = dict(export_augmentation or {})
        root_attrs.update(dict(root_extra_attrs or {}))
        if include_root_workflow:
            root_attrs["workflow"] = wf
        if include_root_total_records:
            root_attrs["total_records"] = str(len(rows))
        root = etree.Element("Records", attrib=root_attrs)

        if include_metadata and self._schema.has_public_metadata:
            public_meta = self.build_public_metadata(session)
            if public_meta:
                meta_el = etree.SubElement(root, "public-metadata")
                for k, v in public_meta.items():
                    fe = etree.SubElement(meta_el, "field", name=str(k))
                    fe.text = "" if v is None else str(v)

        specs = xml_record_attr_specs_resolved or []
        for row in rows:
            flat = self._flatten(row)
            attrs: dict[str, str] = {}
            if include_record_workflow:
                attrs["workflow"] = wf
            if self._schema.record_key_keys and include_record_key:
                attrs["recordKey"] = self._make_record_key_hash(row)
            for spec in specs:
                aname = spec["name"]
                if not should_emit_export_show_attribute(
                    spec.get("show_attribute", True), session
                ):
                    continue
                if spec.get("from_field"):
                    raw = flat.get(spec["from_field"], "")
                    val = "" if raw is None else str(raw)
                else:
                    val = spec.get("static") or ""
                if spec.get("required") and not str(val).strip():
                    raise ValueError(
                        f"[RECORD] required XML <record> attribute '{aname}' is empty for a row"
                    )
                attrs[aname] = val
            attrs.update(dict(record_extra_attrs or {}))
            rec_el = etree.SubElement(root, "record", attrib=attrs)
            for gk, gv in row.items():
                _append_xml_from_value(rec_el, gk, gv)

        tree = etree.ElementTree(root)
        tmp = path.with_name(path.name + ".tmp")
        tree.write(
            str(tmp),
            encoding="utf-8",
            xml_declaration=True,
            pretty_print=True,
        )
        tmp.replace(path)

    def _save_html(
        self,
        path: Path,
        rows: list[dict],
        *,
        include_metadata: bool = False,
        session=None,
        html_title: str | None = None,
        export_augmentation: dict[str, str] | None = None,
        tabular_use_prefix: bool = False,
    ) -> None:
        if not rows:
            return

        flat_rows = self._rows_for_tabular_export(
            rows, tabular_use_prefix=tabular_use_prefix
        )
        keys: list[str] = []
        for row in flat_rows:
            for key in row:
                if key not in keys:
                    keys.append(key)

        title = (html_title or "").strip() or self._workflow_name(session) or "Records"
        title_esc = html_module.escape(title)
        xa = export_augmentation or {}

        parts: list[str] = [
            "<!DOCTYPE html>",
            '<html lang="en">',
            "<head>",
            '<meta charset="utf-8"/>',
            f"<title>{title_esc}</title>",
            "<style>table{border-collapse:collapse;font-family:sans-serif;}th,td{border:1px solid #ccc;padding:4px 8px;}th{background:#f0f0f0;}</style>",
            "</head>",
            "<body>",
            f"<h1>{title_esc}</h1>",
        ]

        if xa:
            parts.append('<section class="francis-export"><h2>Export</h2><table>')
            for ek, ev in xa.items():
                parts.append(
                    "<tr><td>"
                    + html_module.escape(str(ek))
                    + "</td><td>"
                    + html_module.escape(str(ev))
                    + "</td></tr>"
                )
            parts.append("</table></section>")

        if include_metadata and self._schema.has_public_metadata:
            public_meta = self.build_public_metadata(session)
            if public_meta:
                parts.append("<section><h2>Metadata</h2><table>")
                for k, v in public_meta.items():
                    parts.append(
                        "<tr><td>"
                        + html_module.escape(str(k))
                        + "</td><td>"
                        + html_module.escape("" if v is None else str(v))
                        + "</td></tr>"
                    )
                parts.append("</table></section>")

        parts.append("<table><thead><tr>")
        for k in keys:
            parts.append(f"<th>{html_module.escape(k)}</th>")
        parts.append("</tr></thead><tbody>")
        for row in flat_rows:
            parts.append("<tr>")
            for k in keys:
                cell = row.get(k)
                cell_s = "" if cell is None else str(cell)
                parts.append(f"<td>{html_module.escape(cell_s)}</td>")
            parts.append("</tr>")
        parts.append("</tbody></table></body></html>")

        write_text_atomic(path, "\n".join(parts))

    def _save_txt(
        self,
        path: Path,
        rows: list[dict],
        *,
        include_metadata: bool = False,
        export_augmentation: dict[str, str] | None = None,
        tabular_use_prefix: bool = False,
    ) -> None:
        if not rows:
            return

        flat_rows = self._rows_for_tabular_export(
            rows, tabular_use_prefix=tabular_use_prefix
        )
        keys: list[str] = []
        for row in flat_rows:
            for key in row:
                if key not in keys:
                    keys.append(key)

        xa = export_augmentation or {}
        lines: list[str] = []
        if include_metadata and self._schema.has_public_metadata:
            for field in self._schema.public_metadata_fields:
                value = field["value"] or ""
                lines.append(f"# {field['name']}: {value}")
        for ek, ev in xa.items():
            lines.append(f"# {ek}: {ev}")

        lines.append("\t".join(keys))
        for row in flat_rows:
            lines.append(
                "\t".join(
                    "" if row.get(k) is None else str(row.get(k))
                    for k in keys
                )
            )

        write_text_atomic(path, "\n".join(lines) + "\n")

    def _save_excel(
        self,
        path: Path,
        rows: list[dict],
        *,
        include_metadata: bool = False,
        session=None,
        sheet_name: str = "Data",
        metadata_sheet_name: str = "Metadata",
        export_augmentation: dict[str, str] | None = None,
        tabular_use_prefix: bool = False,
    ) -> None:
        if not rows:
            return

        from openpyxl import Workbook

        flat_rows = self._rows_for_tabular_export(
            rows, tabular_use_prefix=tabular_use_prefix
        )
        keys: list[str] = []
        for row in flat_rows:
            for key in row:
                if key not in keys:
                    keys.append(key)

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name[:31] if sheet_name else "Data"

        ws.append(keys)
        for row in flat_rows:
            ws.append(["" if row.get(k) is None else row.get(k) for k in keys])

        xa = export_augmentation or {}
        if include_metadata and self._schema.has_public_metadata:
            public_meta = self.build_public_metadata(session)
            if public_meta:
                ms = wb.create_sheet(metadata_sheet_name[:31] if metadata_sheet_name else "Metadata")
                ms.append(["name", "value"])
                for k, v in public_meta.items():
                    ms.append([k, "" if v is None else v])
        if xa:
            ex = wb.create_sheet("Export"[:31])
            ex.append(["name", "value"])
            for k, v in xa.items():
                ex.append([k, "" if v is None else str(v)])

        tmp = path.with_name(path.name + ".tmp")
        wb.save(tmp)
        tmp.replace(path)

    def _save_parquet(
        self,
        path: Path,
        rows: list[dict],
        export_augmentation: dict[str, str] | None = None,
        *,
        tabular_use_prefix: bool = False,
    ) -> None:
        if not rows:
            return

        import pyarrow as pa
        import pyarrow.parquet as pq

        flat_rows = self._rows_for_tabular_export(
            rows, tabular_use_prefix=tabular_use_prefix
        )
        table = pa.Table.from_pylist(flat_rows)
        xa = export_augmentation or {}
        if xa:
            meta_json = json.dumps(xa, ensure_ascii=False, default=str)
            existing = table.schema.metadata or {}
            merged = {**existing, b"francis_export": meta_json.encode("utf-8")}
            table = table.replace_schema_metadata(merged)
        tmp = path.with_name(path.name + ".tmp")
        pq.write_table(table, tmp, compression="snappy")
        tmp.replace(path)

    def _flatten(self, obj: dict, prefix: str = "") -> dict:
        """
        Nested record row → flat keys like ``listing.price``.
        Preserve ``None`` so JSON/NDJSON exports emit ``null`` (do not coerce to "").
        Formats that need blank cells (CSV, etc.) normalize None at write time.
        """
        result = {}
        for key, value in obj.items():
            full_key = f"{prefix}.{key}" if prefix else key
            if isinstance(value, dict):
                result.update(self._flatten(value, full_key))
            else:
                result[full_key] = value
        return result

    def _flatten_for_export(self, row: dict) -> dict:
        """
        Flatten for file output: drop shared group prefix (e.g. listing.* → short keys)
        and normalize line breaks in string cells.
        """
        flat = self._flatten(row)
        flat = _strip_shared_group_prefix_from_flat(flat)
        return {k: _sanitize_export_scalar_for_flat(v) for k, v in flat.items()}

    def _flatten_sanitize_only(self, row: dict) -> dict:
        """Flat keys keeping group prefix (e.g. listing.field); sanitize string cells."""
        flat = self._flatten(row)
        return {k: _sanitize_export_scalar_for_flat(v) for k, v in flat.items()}

    def _rows_for_tabular_export(
        self, rows: list[dict], *, tabular_use_prefix: bool
    ) -> list[dict]:
        """Tabular formats always use flat columns; prefix mode keeps listing.field names."""
        if tabular_use_prefix:
            return [self._flatten_sanitize_only(r) for r in rows]
        return [self._flatten_for_export(r) for r in rows]

    def _rows_for_json_ndjson_export(self, rows: list[dict], *, shape: str) -> list[dict]:
        """
        shape:
            clean — short keys (default)
            prefix — dotted group keys (listing.field)
            nested — original row dict structure (groups preserved)
        """
        if shape == "nested":
            return [_sanitize_nested_export_value(copy.deepcopy(r)) for r in rows]
        if shape == "prefix":
            return [self._flatten_sanitize_only(r) for r in rows]
        return [self._flatten_for_export(r) for r in rows]
