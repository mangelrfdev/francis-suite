"""
tests/test_pipeline.py

End-to-end test of the full execution pipeline.
Tests that a workflow XML can be parsed and executed correctly.
"""
import json

import respx
import httpx
from francis_suite.core.parser import FParser
from francis_suite.core.runtime import FRuntime
from francis_suite.core.session import SessionStatus
from unittest.mock import patch


def test_log_hand_executes():
    """A workflow with a single <log> tag should complete successfully."""
    xml = """
    <francis-workflow>
        <log>Hello Francis Suite</log>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-log")

    assert session.status == SessionStatus.COMPLETED
    assert session.duration is not None
    assert session.error is None


def test_unknown_tag_fails_session():
    """A workflow with an unknown tag should fail the session."""
    xml = """
    <francis-workflow>
        <tag-que-no-existe/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-unknown")

    assert session.status == SessionStatus.FAILED
    assert session.error is not None

def test_box_def_stores_variable():
    """box-def should execute children and store result in context."""
    xml = """
    <francis-workflow>
        <box-def name="mensaje">
            <log>guardando esto</log>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-box-def")

    assert session.status == SessionStatus.COMPLETED
    variable = session.context.get("mensaje")
    assert not variable.is_empty()
    assert variable.to_string() == "guardando esto"

def test_sleep_executes():
    """sleep should pause execution and return empty."""
    xml = """
    <francis-workflow>
        <sleep ms="0"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sleep")

    assert session.status == SessionStatus.COMPLETED


def test_sleep_invalid_seconds_fails():
    """sleep with invalid ms attribute should fail the session."""
    xml = """
    <francis-workflow>
        <sleep ms="abc"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sleep-invalid")

    assert session.status == SessionStatus.FAILED


def test_sleep_variable_mode():
    """sleep with min/avg/max should execute without error."""
    xml = """
    <francis-workflow>
        <sleep>
            <sleep-min>0</sleep-min>
            <sleep-avg>0</sleep-avg>
            <sleep-max>10</sleep-max>
        </sleep>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sleep-variable")

    assert session.status == SessionStatus.COMPLETED

def test_httpx_call_fetches_url():
    """httpx-call should fetch a URL and return the response body."""
    xml = """
    <francis-workflow>
        <box-def name="page">
            <httpx-call url="https://example.com"/>
        </box-def>
    </francis-workflow>
    """

    with respx.mock:
        respx.get("https://example.com").mock(
            return_value=httpx.Response(200, text="<html>Hello</html>")
        )

        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-call")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("page")
    assert not result.is_empty()
    assert "<html>Hello</html>" in result.to_string()

def test_convert_html_to_xml():
    """convert-html-to-xml should clean HTML and return valid XML."""
    xml = """
    <francis-workflow>
        <box-def name="clean">
            <convert-html-to-xml>
                <log>&lt;html&gt;&lt;body&gt;&lt;h1&gt;Hello&lt;/h1&gt;&lt;/body&gt;&lt;/html&gt;</log>
            </convert-html-to-xml>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-convert")

    print(f"\nERROR: {session.error}")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("clean")
    assert not result.is_empty()
    assert "h1" in result.to_string()

def test_xpath_extract_gets_text():
    """xpath-extract should apply XPath and return matching results."""
    xml_workflow = """
    <francis-workflow>
        <box-def name="resultado">
            <xpath-extract expression="//h1/text()"><![CDATA[<html><body><h1>Hola mundo</h1></body></html>]]></xpath-extract>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml_workflow)
    session = runtime.run(root, workflow_name="test-xpath")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("resultado")
    assert not result.is_empty()
    assert "Hola mundo" in result.to_string()

def test_loop_iterates_over_list():
    """loop should iterate over a list and execute children for each item."""
    xml = """
    <francis-workflow>
        <box-def name="frutas">
            <build-list>
                <log>manzana</log>
                <log>pera</log>
                <log>uva</log>
            </build-list>
        </box-def>
        <loop item="fruta" index="i" max-loops="10">
            <loop-list>
                <box name="frutas"/>
            </loop-list>
            <loop-body>
                <log>Fruta ${i}: ${fruta}</log>
            </loop-body>
        </loop>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-loop")
    assert session.status == SessionStatus.COMPLETED

def test_if_executes_when_true():
    """if should execute children when condition is true."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <if condition="1 == 1">
                <log>condicion verdadera</log>
            </if>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-if-true")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "condicion verdadera"

def test_if_skips_when_false():
    """if should skip children when condition is false."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <if condition="1 == 2">
                <log>no deberia ejecutarse</log>
            </if>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-if-false")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").is_empty()

def test_box_def_stores_variable():
    """box-def should execute children and store result in context."""
    xml = """
    <francis-workflow>
        <box-def name="mensaje">
            <log>guardando esto</log>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-box-def")

    assert session.status == SessionStatus.COMPLETED
    variable = session.context.get("mensaje")
    assert not variable.is_empty()
    assert variable.to_string() == "guardando esto"


def test_box_retrieves_variable():
    """box should retrieve a previously stored variable."""
    xml = """
    <francis-workflow>
        <box-def name="mensaje">
            <log>hola</log>
        </box-def>
        <box-def name="copia">
            <box name="mensaje"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-box")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("copia").to_string() == "hola"

def test_while_executes_while_true():
    """while should execute children while condition is true."""
    xml = """
    <francis-workflow>
        <while condition="1 == 2">
            <log>no deberia ejecutarse</log>
        </while>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-while")

    assert session.status == SessionStatus.COMPLETED

def test_try_completes_when_no_error():
    """try should complete normally when no error occurs."""
    xml = """
    <francis-workflow>
        <try>
            <log>sin error</log>
            <catch>
                <log>no deberia ejecutarse</log>
            </catch>
        </try>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-try-ok")

    assert session.status == SessionStatus.COMPLETED

def test_try_executes_catch_on_error():
    """try should execute catch block when an error occurs."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <try>
                <tag-que-no-existe/>
                <catch>
                    <log>error capturado</log>
                </catch>
            </try>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-try-catch")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "error capturado"

def test_function_create_and_call():
    """function-create should define and function-call should execute it."""
    xml = """
    <francis-workflow>
        <function-create name="saludar">
            <box-def name="msg">
                <log>hola desde funcion</log>
            </box-def>
            <function-return>
                <box name="msg"/>
            </function-return>
        </function-create>
        <box-def name="resultado">
            <function-call name="saludar"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-function")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "hola desde funcion"


def test_function_call_with_params():
    """function-call should inject params into function scope."""
    xml = """
    <francis-workflow>
        <function-create name="repetir">
            <function-return>
                <box name="valor"/>
            </function-return>
        </function-create>
        <box-def name="resultado">
            <function-call name="repetir">
                <function-param name="valor">texto de prueba</function-param>
            </function-call>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-function-params")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "texto de prueba"


def test_function_call_undefined_fails():
    """function-call with undefined function should fail the session."""
    xml = """
    <francis-workflow>
        <function-call name="no-existe"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-function-undefined")

    assert session.status == SessionStatus.FAILED

def test_regex_finds_match():
    """regex should apply pattern and return match."""
    xml = r"""
    <francis-workflow>
        <box-def name="precio">
            <regex>
                <regex-pattern><![CDATA[\d+\.\d{2}]]></regex-pattern>
                <regex-input>El precio es 19.99 euros</regex-input>
            </regex>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-regex")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("precio").to_string() == "19.99"


def test_regex_with_groups_and_template():
    """regex should apply template with capture groups."""
    xml = r"""
    <francis-workflow>
        <box-def name="telefono">
            <regex>
                <regex-pattern><![CDATA[(\d{3})-(\d{3})-(\d{4})]]></regex-pattern>
                <regex-input>Llama al 555-123-4567</regex-input>
                <regex-result>(${_1}) ${_2}-${_3}</regex-result>
            </regex>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-regex-groups")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("telefono").to_string() == "(555) 123-4567"


def test_regex_no_match_returns_empty():
    """regex with no match should return empty."""
    xml = r"""
    <francis-workflow>
        <box-def name="resultado">
            <regex>
                <regex-pattern><![CDATA[\d+]]></regex-pattern>
                <regex-input>no hay numeros aqui</regex-input>
            </regex>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-regex-empty")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").is_empty()

def test_compose_interpolates_variables():
    """compose should replace ${var} with context values."""
    xml = """
    <francis-workflow>
        <box-def name="nombre">
            <log>Francis</log>
        </box-def>
        <box-def name="mensaje">
            <compose>Hola ${nombre}!</compose>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-compose")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("mensaje").to_string() == "Hola Francis!"

def test_compose_unknown_var_stays():
    """compose should return empty string for unknown variables."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <compose>Valor: ${no-existe}</compose>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-compose-unknown")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "Valor: "

def test_text_split_splits_by_delimiter():
    """text-split should split text by delimiter and return a list."""
    xml = """
    <francis-workflow>
        <box-def name="frutas">
            <text-split delimiter=",">manzana,pera,uva</text-split>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-text-split")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("frutas")
    assert not result.is_empty()
    assert len(result.to_list()) == 3

def test_text_split_trims_tokens():
    """text-split should trim whitespace from tokens by default."""
    xml = """
    <francis-workflow>
        <box-def name="items">
            <text-split delimiter=",">  uno  ,  dos  ,  tres  </text-split>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-text-split-trim")

    assert session.status == SessionStatus.COMPLETED
    items = session.context.get("items").to_list()
    assert items[0].to_string() == "uno"
    assert items[1].to_string() == "dos"
    assert items[2].to_string() == "tres"

def test_evaluate_arithmetic():
    """evaluate should compute arithmetic expressions."""
    xml = """
    <francis-workflow>
        <box-def name="precio">
            <log>10</log>
        </box-def>
        <box-def name="cantidad">
            <log>3</log>
        </box-def>
        <box-def name="total">
            <evaluate>${precio} * ${cantidad}</evaluate>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-evaluate-arithmetic")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("total").to_string() == "30"

def test_evaluate_is_empty():
    """evaluate should support isEmpty() method call."""
    xml = """
    <francis-workflow>
        <box-def name="nombre">
            <log>Francis</log>
        </box-def>
        <box-def name="resultado">
            <evaluate>${nombre.isEmpty()}</evaluate>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-evaluate-isempty")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "False"

def test_evaluate_to_upper():
    """evaluate should support toUpperCase() method call."""
    xml = """
    <francis-workflow>
        <box-def name="nombre">
            <log>francis</log>
        </box-def>
        <box-def name="resultado">
            <evaluate>${nombre.toUpperCase()}</evaluate>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-evaluate-upper")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "FRANCIS"

def test_else_executes_when_if_false():
    """else should execute when preceding if condition is false."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <log>nada</log>
        </box-def>
        <if condition="1 == 2">
            <box-def name="resultado">
                <log>if ejecutado</log>
            </box-def>
        </if>
        <else>
            <box-def name="resultado">
                <log>else ejecutado</log>
            </box-def>
        </else>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-else")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "else ejecutado"

def test_else_skips_when_if_true():
    """else should not execute when preceding if condition is true."""
    xml = """
    <francis-workflow>
        <if condition="1 == 1">
            <box-def name="resultado">
                <log>if ejecutado</log>
            </box-def>
        </if>
        <else>
            <box-def name="resultado">
                <log>else ejecutado</log>
            </box-def>
        </else>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-else-skip")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "if ejecutado"

def test_case_executes_first_match():
    """case should execute only the first matching if."""
    xml = """
    <francis-workflow>
        <box-def name="tipo">
            <log>B</log>
        </box-def>
        <box-def name="resultado">
            <case>
                <if condition="${tipo} == 'A'">
                    <log>es tipo A</log>
                </if>
                <if condition="${tipo} == 'B'">
                    <log>es tipo B</log>
                </if>
                <if condition="${tipo} == 'C'">
                    <log>es tipo C</log>
                </if>
                <else>
                    <log>tipo desconocido</log>
                </else>
            </case>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-case")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "es tipo B"

def test_case_executes_else_when_no_match():
    """case should execute else when no if matches."""
    xml = """
    <francis-workflow>
        <box-def name="tipo">
            <log>Z</log>
        </box-def>
        <box-def name="resultado">
            <case>
                <if condition="${tipo} == 'A'">
                    <log>es tipo A</log>
                </if>
                <if condition="${tipo} == 'B'">
                    <log>es tipo B</log>
                </if>
                <else>
                    <log>tipo desconocido</log>
                </else>
            </case>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-case-else")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "tipo desconocido"

def test_exit_stops_workflow():
    """exit should stop workflow execution cleanly."""
    xml = """
    <francis-workflow>
        <box-def name="antes">
            <log>antes del exit</log>
        </box-def>
        <exit/>
        <box-def name="despues">
            <log>despues del exit</log>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-exit")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("antes").to_string() == "antes del exit"
    assert session.context.get("despues").is_empty()

def test_build_list_creates_list():
    """build-list should create a FListVariable from children results."""
    xml = """
    <francis-workflow>
        <box-def name="items">
            <build-list>
                <log>uno</log>
                <log>dos</log>
                <log>tres</log>
            </build-list>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-build-list")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("items")
    assert not result.is_empty()
    assert len(result.to_list()) == 3
    assert result.to_list()[0].to_string() == "uno"
    assert result.to_list()[1].to_string() == "dos"
    assert result.to_list()[2].to_string() == "tres"

def test_call_workflow_executes_external_file(tmp_path):
    """call-workflow should load and execute an external workflow file."""
    external = tmp_path / "external.xml"
    external.write_text("""
    <francis-workflow>
        <box-def name="externo">
            <log>valor externo</log>
        </box-def>
    </francis-workflow>
    """)

    xml = f"""
    <francis-workflow>
        <call-workflow path="{external.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-call-workflow")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("externo").to_string() == "valor externo"

def test_convert_json_to_xml():
    """convert-json-to-xml should convert JSON string to XML."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-json-to-xml root="persona">{"nombre": "Francis", "edad": "30"}</convert-json-to-xml>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-json-to-xml")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("resultado").to_string()
    assert "<nombre>Francis</nombre>" in result
    assert "<edad>30</edad>" in result

def test_convert_xml_to_json():
    """convert-xml-to-json should convert XML string to JSON."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-xml-to-json>&lt;persona&gt;&lt;nombre&gt;Francis&lt;/nombre&gt;&lt;/persona&gt;</convert-xml-to-json>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-to-json")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("resultado").to_string()
    assert "Francis" in result
    assert "nombre" in result

def test_file_write_and_read(tmp_path):
    """file-write should write content and file-read should read it back."""
    output = tmp_path / "test.txt"

    xml = f"""
    <francis-workflow>
        <file-write path="{output.as_posix()}">Hola Francis</file-write>
        <box-def name="contenido">
            <file-read path="{output.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("contenido").to_string() == "Hola Francis"

def test_file_write_append(tmp_path):
    """file-write with append=true should append content."""
    output = tmp_path / "test.txt"

    xml = f"""
    <francis-workflow>
        <file-write path="{output.as_posix()}">linea 1</file-write>
        <file-write path="{output.as_posix()}" append="true"> linea 2</file-write>
        <box-def name="contenido">
            <file-read path="{output.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-append")

    assert session.status == SessionStatus.COMPLETED
    assert "linea 1" in session.context.get("contenido").to_string()
    assert "linea 2" in session.context.get("contenido").to_string()

def test_file_manage_delete(tmp_path):
    """file-manage delete should remove a file."""
    target = tmp_path / "delete_me.txt"
    target.write_text("borrame")

    xml = f"""
    <francis-workflow>
        <file-manage action="delete" path="{target.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-delete")

    assert session.status == SessionStatus.COMPLETED
    assert not target.exists()

def test_file_manage_list(tmp_path):
    """file-manage list should return list of files."""
    (tmp_path / "a.txt").write_text("a")
    (tmp_path / "b.txt").write_text("b")
    (tmp_path / "c.txt").write_text("c")

    xml = f"""
    <francis-workflow>
        <box-def name="archivos">
            <file-manage action="list" path="{tmp_path.as_posix()}" filter="*.txt"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-list")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("archivos")
    assert len(result.to_list()) == 3

def test_file_manage_copy(tmp_path):
    """file-manage copy should copy a file to destination."""
    source = tmp_path / "original.txt"
    dest = tmp_path / "copia.txt"
    source.write_text("contenido original")

    xml = f"""
    <francis-workflow>
        <file-manage action="copy" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-copy")

    assert session.status == SessionStatus.COMPLETED
    assert dest.exists()
    assert dest.read_text() == "contenido original"

def test_file_manage_move(tmp_path):
    """file-manage move should move a file to destination."""
    source = tmp_path / "original.txt"
    dest = tmp_path / "movido.txt"
    source.write_text("contenido")

    xml = f"""
    <francis-workflow>
        <file-manage action="move" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-move")

    assert session.status == SessionStatus.COMPLETED
    assert dest.exists()
    assert not source.exists()
    assert dest.read_text() == "contenido"

def test_file_read_not_found():
    """file-read should fail when file does not exist."""
    xml = """
    <francis-workflow>
        <file-read path="no_existe.txt"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-read-not-found")

    assert session.status == SessionStatus.FAILED

def test_shared_box_def_stores_in_global_scope():
    """shared-box-def should store variable in global scope."""
    xml = """
    <francis-workflow>
        <shared-box-def name="env">production</shared-box-def>
        <box-def name="resultado">
            <shared-box name="env"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-shared-box-def")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "production"

def test_shared_box_def_replace_false_does_not_overwrite():
    """shared-box-def with replace=false should not overwrite existing value."""
    xml = """
    <francis-workflow>
        <shared-box-def name="env" replace="false">production</shared-box-def>
        <shared-box-def name="env" replace="false">staging</shared-box-def>
        <box-def name="resultado">
            <shared-box name="env"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-shared-box-no-replace")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "production"

def test_shared_box_def_replace_true_overwrites():
    """shared-box-def with replace=true should overwrite existing value."""
    xml = """
    <francis-workflow>
        <shared-box-def name="env" replace="false">production</shared-box-def>
        <shared-box-def name="env" replace="true">staging</shared-box-def>
        <box-def name="resultado">
            <shared-box name="env"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-shared-box-replace")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "staging"


def test_shared_box_accessible_as_variable():
    """shared-box-def variable should be accessible via ${variable} syntax."""
    xml = """
    <francis-workflow>
        <shared-box-def name="env">production</shared-box-def>
        <box-def name="resultado">
            <compose>entorno: ${env}</compose>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-shared-box-variable")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "entorno: production"

def test_shared_box_accessible_inside_function():
    """shared-box-def should be accessible inside functions."""
    xml = """
    <francis-workflow>
        <shared-box-def name="env">production</shared-box-def>
        <function-create name="get-env">
            <function-return>
                <shared-box name="env"/>
            </function-return>
        </function-create>
        <box-def name="resultado">
            <function-call name="get-env"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-shared-box-in-function")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "production"

def test_shared_box_used_in_condition():
    """shared-box-def should be usable in if conditions."""
    xml = """
    <francis-workflow>
        <shared-box-def name="activo">true</shared-box-def>
        <box-def name="resultado">
            <if condition="${activo.toBoolean()}">
                <log>activo</log>
            </if>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-shared-box-condition")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "activo"

def test_function_create_replace_false_does_not_overwrite():
    """function-create with replace=false should not overwrite existing function."""
    xml = """
    <francis-workflow>
        <function-create name="saludar" replace="false">
            <function-return>
                <log>hola original</log>
            </function-return>
        </function-create>
        <function-create name="saludar" replace="false">
            <function-return>
                <log>hola reemplazado</log>
            </function-return>
        </function-create>
        <box-def name="resultado">
            <function-call name="saludar"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-function-no-replace")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "hola original"

def test_function_create_replace_true_overwrites():
    """function-create with replace=true should overwrite existing function."""
    xml = """
    <francis-workflow>
        <function-create name="saludar" replace="false">
            <function-return>
                <log>hola original</log>
            </function-return>
        </function-create>
        <function-create name="saludar" replace="true">
            <function-return>
                <log>hola reemplazado</log>
            </function-return>
        </function-create>
        <box-def name="resultado">
            <function-call name="saludar"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-function-replace")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "hola reemplazado"

def test_box_def_resolves_variables_in_body_text():
    """box-def should resolve ${variables} directly in body text without compose."""
    xml = """
    <francis-workflow>
        <box-def name="base_url">ejemplo.com</box-def>
        <box-def name="pagina">3</box-def>
        <box-def name="url">https://${base_url}/page-${pagina}.html</box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-box-def-body-text")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("url").to_string() == "https://ejemplo.com/page-3.html"

def test_file_write_newline(tmp_path):
    """file-write with newline=true should append a newline after content."""
    output = tmp_path / "output.txt"

    xml = f"""
    <francis-workflow>
        <file-write path="{output.as_posix()}" newline="true">linea uno</file-write>
        <file-write path="{output.as_posix()}" append="true" newline="true">linea dos</file-write>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-file-write-newline")

    assert session.status == SessionStatus.COMPLETED
    assert output.read_text() == "linea uno\nlinea dos\n"

def test_httpx_call_binary_response(tmp_path):
    """httpx-call with response=binary should return bytes."""
    fake_pdf = b"%PDF-1.4 fake pdf content"

    xml = """
    <francis-workflow>
        <box-def name="reporte">
            <httpx-call url="https://example.com/report.pdf" response="binary"/>
        </box-def>
    </francis-workflow>
    """

    with respx.mock:
        respx.get("https://example.com/report.pdf").mock(
            return_value=httpx.Response(200, content=fake_pdf)
        )

        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-binary")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("reporte")
    assert not result.is_empty()
    assert result.value == fake_pdf


def test_httpx_call_binary_write_to_disk(tmp_path):
    """httpx-call binary + file-write binary should save file correctly."""
    fake_pdf = b"%PDF-1.4 fake pdf content"
    output = tmp_path / "report.pdf"

    xml = f"""
    <francis-workflow>
        <box-def name="reporte">
            <httpx-call url="https://example.com/report.pdf" response="binary"/>
        </box-def>
        <file-write path="{output.as_posix()}" encoding="binary">
            <box name="reporte"/>
        </file-write>
    </francis-workflow>
    """

    with respx.mock:
        respx.get("https://example.com/report.pdf").mock(
            return_value=httpx.Response(200, content=fake_pdf)
        )

        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-binary-write")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    assert output.read_bytes() == fake_pdf


def test_httpx_call_stream_saves_file(tmp_path):
    """httpx-call stream should save file to disk and return path."""
    fake_content = b"fake video content chunk"
    output = tmp_path / "video.mp4"

    xml = f"""
    <francis-workflow>
        <box-def name="archivo">
            <httpx-call url="https://example.com/video.mp4" response="stream" path="{output.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    with respx.mock:
        respx.get("https://example.com/video.mp4").mock(
            return_value=httpx.Response(200, content=fake_content)
        )

        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-stream")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    assert output.read_bytes() == fake_content

    # tmp file should not exist after successful download
    tmp_file = output.with_suffix(output.suffix + ".tmp")
    assert not tmp_file.exists()

    # box should contain the path where file was saved
    result = session.context.get("archivo")
    assert not result.is_empty()
    assert result.to_string() == output.as_posix()


def test_httpx_call_stream_cleans_tmp_on_failure(tmp_path):
    """httpx-call stream should remove .tmp file if download fails."""
    output = tmp_path / "video.mp4"
    tmp_file = output.with_suffix(output.suffix + ".tmp")

    xml = f"""
    <francis-workflow>
        <httpx-call url="https://example.com/video.mp4" response="stream" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    with respx.mock:
        respx.get("https://example.com/video.mp4").mock(
            return_value=httpx.Response(500)
        )

        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-stream-fail")

    assert session.status == SessionStatus.FAILED
    assert not output.exists()
    assert not tmp_file.exists()


def test_httpx_call_invalid_response_fails():
    """httpx-call with invalid response attribute should fail."""
    xml = """
    <francis-workflow>
        <httpx-call url="https://example.com" response="invalid"/>
    </francis-workflow>
    """

    with respx.mock:
        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-invalid-response")

    assert session.status == SessionStatus.FAILED


def test_httpx_call_stream_missing_path_fails():
    """httpx-call stream without path attribute should fail."""
    xml = """
    <francis-workflow>
        <httpx-call url="https://example.com/video.mp4" response="stream"/>
    </francis-workflow>
    """

    with respx.mock:
        parser = FParser()
        runtime = FRuntime()

        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-httpx-stream-no-path")

    assert session.status == SessionStatus.FAILED

def test_file_manage_mkdir_creates_directory(tmp_path):
    """file-manage mkdir should create a directory."""
    new_dir = tmp_path / "nueva_carpeta"

    xml = f"""
    <francis-workflow>
        <file-manage action="mkdir" path="{new_dir.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-mkdir")

    assert session.status == SessionStatus.COMPLETED
    assert new_dir.exists()
    assert new_dir.is_dir()


def test_file_manage_mkdir_skips_if_exists(tmp_path):
    """file-manage mkdir should not fail if directory already exists."""
    existing_dir = tmp_path / "existente"
    existing_dir.mkdir()

    xml = f"""
    <francis-workflow>
        <file-manage action="mkdir" path="{existing_dir.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-mkdir-exists")

    assert session.status == SessionStatus.COMPLETED
    assert existing_dir.exists()


def test_file_manage_mkdir_creates_nested_directories(tmp_path):
    """file-manage mkdir should create all parent directories."""
    nested = tmp_path / "nivel1" / "nivel2" / "nivel3"

    xml = f"""
    <francis-workflow>
        <file-manage action="mkdir" path="{nested.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-mkdir-nested")

    assert session.status == SessionStatus.COMPLETED
    assert nested.exists()
    assert nested.is_dir()


def test_file_manage_check_exists_true(tmp_path):
    """file-manage check-exists should return true when file exists."""
    archivo = tmp_path / "reporte.pdf"
    archivo.write_text("contenido")

    xml = f"""
    <francis-workflow>
        <box-def name="existe">
            <file-manage action="check-exists" path="{archivo.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-check-exists-true")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("existe").to_string() == "true"


def test_file_manage_check_exists_false(tmp_path):
    """file-manage check-exists should return false when file does not exist."""
    archivo = tmp_path / "no_existe.pdf"

    xml = f"""
    <francis-workflow>
        <box-def name="existe">
            <file-manage action="check-exists" path="{archivo.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-check-exists-false")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("existe").to_string() == "false"


def test_file_manage_check_exists_directory(tmp_path):
    """file-manage check-exists should return true for directories."""
    carpeta = tmp_path / "fotos"
    carpeta.mkdir()

    xml = f"""
    <francis-workflow>
        <box-def name="existe">
            <file-manage action="check-exists" path="{carpeta.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-check-exists-dir")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("existe").to_string() == "true"


def test_file_manage_get_size_file(tmp_path):
    """file-manage get-size should return file size in bytes."""
    archivo = tmp_path / "reporte.txt"
    archivo.write_text("contenido de prueba", encoding="utf-8")

    xml = f"""
    <francis-workflow>
        <box-def name="tamano">
            <file-manage action="get-size" path="{archivo.as_posix()}"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-get-size-file")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("tamano").to_string()
    assert "bytes" in result


def test_file_manage_get_size_auto_format(tmp_path):
    """file-manage get-size with size-format=auto should return formatted size."""
    archivo = tmp_path / "reporte.txt"
    archivo.write_text("contenido de prueba", encoding="utf-8")

    xml = f"""
    <francis-workflow>
        <box-def name="tamano">
            <file-manage action="get-size" path="{archivo.as_posix()}" size-format="auto"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-get-size-auto")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("tamano").to_string()
    assert result != ""


def test_file_manage_get_size_not_found_fails(tmp_path):
    """file-manage get-size should fail when file does not exist."""
    xml = f"""
    <francis-workflow>
        <file-manage action="get-size" path="{(tmp_path / 'no_existe.txt').as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-get-size-not-found")

    assert session.status == SessionStatus.FAILED


def test_file_manage_rename_file(tmp_path):
    """file-manage rename should rename a file."""
    source = tmp_path / "foto_1.jpg"
    dest   = tmp_path / "foto_001.jpg"
    source.write_text("contenido foto")

    xml = f"""
    <francis-workflow>
        <file-manage action="rename" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-rename")

    assert session.status == SessionStatus.COMPLETED
    assert dest.exists()
    assert not source.exists()
    assert dest.read_text() == "contenido foto"


def test_file_manage_rename_not_found_fails(tmp_path):
    """file-manage rename should fail when source does not exist."""
    xml = f"""
    <francis-workflow>
        <file-manage action="rename"
            path="{(tmp_path / 'no_existe.jpg').as_posix()}"
            to="{(tmp_path / 'nuevo.jpg').as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-rename-not-found")

    assert session.status == SessionStatus.FAILED


def test_file_manage_rename_destination_exists_fails(tmp_path):
    """file-manage rename should fail when destination already exists."""
    source = tmp_path / "foto_1.jpg"
    dest   = tmp_path / "foto_001.jpg"
    source.write_text("original")
    dest.write_text("ya existe")

    xml = f"""
    <francis-workflow>
        <file-manage action="rename" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-rename-dest-exists")

    assert session.status == SessionStatus.FAILED


def test_file_manage_rename_different_directory_fails(tmp_path):
    """file-manage rename should fail when to is in a different directory."""
    source  = tmp_path / "foto.jpg"
    subdir  = tmp_path / "sub"
    subdir.mkdir()
    dest    = subdir / "foto.jpg"
    source.write_text("contenido")

    xml = f"""
    <francis-workflow>
        <file-manage action="rename" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-rename-diff-dir")

    assert session.status == SessionStatus.FAILED


def test_file_manage_delete_non_empty_fails(tmp_path):
    """file-manage delete should fail on non-empty directory without force-delete."""
    carpeta = tmp_path / "fotos"
    carpeta.mkdir()
    (carpeta / "foto.jpg").write_text("foto")

    xml = f"""
    <francis-workflow>
        <file-manage action="delete" path="{carpeta.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-delete-non-empty")

    assert session.status == SessionStatus.FAILED
    assert carpeta.exists()


def test_file_manage_delete_force_deletes_non_empty(tmp_path):
    """file-manage delete with force-delete=true should delete non-empty directory."""
    carpeta = tmp_path / "fotos"
    carpeta.mkdir()
    (carpeta / "foto.jpg").write_text("foto")

    xml = f"""
    <francis-workflow>
        <file-manage action="delete" path="{carpeta.as_posix()}" force-delete="true"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-delete-force")

    assert session.status == SessionStatus.COMPLETED
    assert not carpeta.exists()


def test_file_manage_copy_destination_exists_fails(tmp_path):
    """file-manage copy should fail when destination exists without force-copy."""
    source = tmp_path / "original.txt"
    dest   = tmp_path / "copia.txt"
    source.write_text("original")
    dest.write_text("ya existe")

    xml = f"""
    <francis-workflow>
        <file-manage action="copy" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-copy-dest-exists")

    assert session.status == SessionStatus.FAILED


def test_file_manage_copy_force_overwrites(tmp_path):
    """file-manage copy with force-copy=true should overwrite destination."""
    source = tmp_path / "original.txt"
    dest   = tmp_path / "copia.txt"
    source.write_text("contenido nuevo")
    dest.write_text("contenido viejo")

    xml = f"""
    <francis-workflow>
        <file-manage action="copy" path="{source.as_posix()}" to="{dest.as_posix()}" force-copy="true"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-copy-force")

    assert session.status == SessionStatus.COMPLETED
    assert dest.read_text() == "contenido nuevo"


def test_file_manage_move_destination_exists_fails(tmp_path):
    """file-manage move should fail when destination exists without force-move."""
    source = tmp_path / "original.txt"
    dest   = tmp_path / "movido.txt"
    source.write_text("original")
    dest.write_text("ya existe")

    xml = f"""
    <francis-workflow>
        <file-manage action="move" path="{source.as_posix()}" to="{dest.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-move-dest-exists")

    assert session.status == SessionStatus.FAILED


def test_file_manage_move_force_overwrites(tmp_path):
    """file-manage move with force-move=true should overwrite destination."""
    source = tmp_path / "original.txt"
    dest   = tmp_path / "movido.txt"
    source.write_text("contenido nuevo")
    dest.write_text("contenido viejo")

    xml = f"""
    <francis-workflow>
        <file-manage action="move" path="{source.as_posix()}" to="{dest.as_posix()}" force-move="true"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-move-force")

    assert session.status == SessionStatus.COMPLETED
    assert dest.read_text() == "contenido nuevo"
    assert not source.exists()


def test_file_manage_list_type_folders(tmp_path):
    """file-manage list with type=folders should return only directories."""
    (tmp_path / "foto.jpg").write_text("foto")
    (tmp_path / "subcarpeta").mkdir()

    xml = f"""
    <francis-workflow>
        <box-def name="carpetas">
            <file-manage action="list" path="{tmp_path.as_posix()}" type="folders"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-list-folders")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("carpetas")
    items  = result.to_list()
    assert len(items) == 1
    assert items[0].to_string().endswith("/")


def test_file_manage_list_type_all(tmp_path):
    """file-manage list with type=all should return files and directories."""
    (tmp_path / "foto.jpg").write_text("foto")
    (tmp_path / "subcarpeta").mkdir()

    xml = f"""
    <francis-workflow>
        <box-def name="todo">
            <file-manage action="list" path="{tmp_path.as_posix()}" type="all"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-list-all")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("todo")
    assert len(result.to_list()) == 2


def test_file_manage_list_search_in_subfolders(tmp_path):
    """file-manage list with search-in-subfolders=true should find files in subdirectories."""
    (tmp_path / "foto1.jpg").write_text("foto1")
    subdir = tmp_path / "sub"
    subdir.mkdir()
    (subdir / "foto2.jpg").write_text("foto2")

    xml = f"""
    <francis-workflow>
        <box-def name="fotos">
            <file-manage action="list" path="{tmp_path.as_posix()}" filter="*.jpg" search-in-subfolders="true"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-list-subfolders")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("fotos")
    assert len(result.to_list()) == 2

def test_sensitive_auto_by_name():
    """Variables with sensitive names should be masked in display."""
    xml = """
    <francis-workflow>
        <shared-box-def name="api_key">sk-abc123xyz</shared-box-def>
        <box-def name="resultado">
            <log>Key: ${api_key}</log>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sensitive-auto")

    assert session.status == SessionStatus.COMPLETED
    # to_string() returns real value — engine uses it internally
    assert session.context.get("api_key").to_string() == "sk-abc123xyz"
    # to_display() returns masked value — logs and UI use it
    assert session.context.get("api_key").to_display() == "*******xyz"


def test_sensitive_explicit_true():
    """Variables with sensitive=true should be masked in display."""
    xml = """
    <francis-workflow>
        <box-def name="codigo_cliente" sensitive="true">abc123</box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sensitive-explicit")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("codigo_cliente").to_string() == "abc123"
    assert session.context.get("codigo_cliente").to_display() == "*******123"


def test_sensitive_explicit_false():
    """Variables with sensitive=false should never be masked."""
    xml = """
    <francis-workflow>
        <box-def name="token_count" sensitive="false">100</box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sensitive-false")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("token_count").to_string() == "100"
    assert session.context.get("token_count").to_display() == "100"


def test_sensitive_short_value():
    """Sensitive variables with short values should show 10 asterisks."""
    xml = """
    <francis-workflow>
        <box-def name="api_key">ab</box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sensitive-short")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("api_key").to_string() == "ab"
    assert session.context.get("api_key").to_display() == "**********"


def test_sensitive_shared_box_def():
    """shared-box-def should also support sensitive flag."""
    xml = """
    <francis-workflow>
        <shared-box-def name="password">mi_clave_secreta</shared-box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-sensitive-shared")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("password").to_string() == "mi_clave_secreta"
    assert "***" in session.context.get("password").to_display()
    assert "mi_clave_secreta" not in session.context.get("password").to_display()

def test_convert_binary_to_base64():
    """convert-binary-to-base64 should encode bytes to base64 string."""
    xml = """
    <francis-workflow>
        <box-def name="texto">hola mundo</box-def>
        <box-def name="resultado">
            <convert-binary-to-base64>${texto}</convert-binary-to-base64>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-binary-to-base64")

    assert session.status == SessionStatus.COMPLETED
    import base64
    expected = base64.b64encode("hola mundo".encode("utf-8")).decode("utf-8")
    assert session.context.get("resultado").to_string() == expected


def test_convert_base64_to_binary():
    """convert-base64-to-binary should decode base64 string to bytes."""
    import base64
    encoded = base64.b64encode(b"hola mundo").decode("utf-8")

    xml = f"""
    <francis-workflow>
        <box-def name="base64">{encoded}</box-def>
        <box-def name="resultado">
            <convert-base64-to-binary>${{base64}}</convert-base64-to-binary>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-base64-to-binary")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").value == b"hola mundo"


def test_convert_base64_to_binary_invalid_fails():
    """convert-base64-to-binary should fail with invalid base64."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-base64-to-binary>esto no es base64!!!</convert-base64-to-binary>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-base64-to-binary-invalid")

    assert session.status == SessionStatus.FAILED


def test_convert_text_to_base64():
    """convert-text-to-base64 should encode text to base64."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-text-to-base64>hola mundo</convert-text-to-base64>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-text-to-base64")

    assert session.status == SessionStatus.COMPLETED
    import base64
    expected = base64.b64encode("hola mundo".encode("utf-8")).decode("utf-8")
    assert session.context.get("resultado").to_string() == expected


def test_convert_base64_to_text():
    """convert-base64-to-text should decode base64 to text."""
    import base64
    encoded = base64.b64encode("hola mundo".encode("utf-8")).decode("utf-8")

    xml = f"""
    <francis-workflow>
        <box-def name="resultado">
            <convert-base64-to-text>{encoded}</convert-base64-to-text>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-base64-to-text")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "hola mundo"


def test_convert_json_to_csv():
    """convert-json-to-csv should convert JSON array to CSV."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-json-to-csv>[{"nombre": "Casa", "precio": "100000"}, {"nombre": "Depto", "precio": "80000"}]</convert-json-to-csv>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-json-to-csv")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("resultado").to_string()
    assert "nombre,precio" in result
    assert "Casa" in result
    assert "Depto" in result


def test_convert_csv_to_json():
    """convert-csv-to-json should convert CSV to JSON array."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-csv-to-json>nombre,precio
Casa,100000
Depto,80000</convert-csv-to-json>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-csv-to-json")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("resultado").to_string()
    assert "Casa" in result
    assert "100000" in result


def test_convert_xml_to_csv():
    """convert-xml-to-csv should convert XML to CSV."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-xml-to-csv>&lt;items&gt;&lt;item&gt;&lt;nombre&gt;Casa&lt;/nombre&gt;&lt;precio&gt;100000&lt;/precio&gt;&lt;/item&gt;&lt;/items&gt;</convert-xml-to-csv>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-to-csv")

    assert session.status == SessionStatus.COMPLETED
    result = session.context.get("resultado").to_string()
    assert "nombre" in result
    assert "Casa" in result


def test_convert_text_to_url():
    """convert-text-to-url should encode text for URL use."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-text-to-url>departamento en santiago</convert-text-to-url>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-text-to-url")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "departamento%20en%20santiago"


def test_convert_url_to_text():
    """convert-url-to-text should decode URL-encoded string."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-url-to-text>departamento%20en%20santiago</convert-url-to-text>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-url-to-text")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "departamento en santiago"


def test_convert_html_entities_to_text():
    """convert-html-entities-to-text should decode HTML entities."""
    xml = """
    <francis-workflow>
        <box-def name="resultado">
            <convert-html-entities-to-text>Casa &amp;amp; Jard&amp;iacute;n</convert-html-entities-to-text>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-html-entities-to-text")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("resultado").to_string() == "Casa & Jardín"

from unittest.mock import patch

def test_pause_task_dev_pauses(monkeypatch):
    """pause-task in dev should pause and wait for input."""
    monkeypatch.setenv("FRANCIS_ENV", "dev")

    xml = """
    <francis-workflow>
        <pause-task message="Revisando datos"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    with patch("builtins.input", return_value=""):
        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-pause-dev")

    assert session.status == SessionStatus.COMPLETED


def test_pause_task_prod_does_not_pause(monkeypatch):
    """pause-task in prod should warn and continue without pausing."""
    monkeypatch.setenv("FRANCIS_ENV", "prod")

    xml = """
    <francis-workflow>
        <pause-task message="Revisando datos"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-pause-prod")

    assert session.status == SessionStatus.COMPLETED


def test_pause_task_no_message(monkeypatch):
    """pause-task without message should work fine."""
    monkeypatch.setenv("FRANCIS_ENV", "dev")

    xml = """
    <francis-workflow>
        <pause-task/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    with patch("builtins.input", return_value=""):
        root = parser.parse_string(xml)
        session = runtime.run(root, workflow_name="test-pause-no-message")

    assert session.status == SessionStatus.COMPLETED

def test_record_create_basic():
    """record-create should create a record in global context."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string"  required="true"/>
                <record-set-field name="precio" type="integer" required="true"/>
            </record-set-group>
        </record-create>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-create")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert isinstance(record, FRecord)
    assert record.count == 0


def test_record_add_adds_row():
    """record-add should add a normalized row to the record."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string"  required="true"/>
                <record-set-field name="precio" type="integer" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
                <record-add-field name="precio">100000</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-add")
    
    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.count == 1
    assert record.last_row["item"]["nombre"] == "Casa"
    assert record.last_row["item"]["precio"] == 100000


def test_record_add_integer_cleans_currency():
    """record-add should clean currency symbols from integer fields."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="precio" type="integer" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="precio">$3.990</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-integer-clean")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.last_row["item"]["precio"] == 3990


def test_record_add_decimal_precision():
    """record-add should normalize decimal with precision."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="precio" type="decimal" precision="2" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="precio">3990.5678</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-decimal")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.last_row["item"]["precio"] == "3990.57"


def test_record_add_boolean_normalizes():
    """record-add should normalize boolean values."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="activo" type="boolean" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="activo">si</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-boolean")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.last_row["item"]["activo"] == True


def test_record_add_null_if_empty():
    """record-add should store None when null-if-empty=true and value is empty."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string"  required="true"/>
                <record-set-field name="marca"  type="string"  required="false" null-if-empty="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
                <record-add-field name="marca"></record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-null-if-empty")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.last_row["item"]["marca"] is None


def test_record_add_uuid_generates_if_empty():
    """record-add should generate UUID when field is empty and type is uuid."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="id"     type="uuid"   required="true"/>
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="id"></record-add-field>
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-uuid")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    generated_id = record.last_row["item"]["id"]
    assert generated_id != ""
    assert len(generated_id) == 36  # UUID format: xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx


def test_record_count_returns_row_count():
    """record-count should return the number of rows."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Depto</record-add-field>
            </record-add-group>
        </record-add>
        <box-def name="total">
            <record-count from="testRecords"/>
        </box-def>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-count")

    assert session.status == SessionStatus.COMPLETED
    assert session.context.get("total").to_string() == "2"


def test_record_last_added_shows_last_row():
    """record-last-added should not fail and show last row."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-last-added from="testRecords"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-last-added")

    assert session.status == SessionStatus.COMPLETED


def test_record_save_json(tmp_path):
    """record-save should save record as JSON file."""
    output = tmp_path / "test.json"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string"  required="true"/>
                <record-set-field name="precio" type="integer" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
                <record-add-field name="precio">100000</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="json" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-json")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()

    import json
    data = json.loads(output.read_text(encoding="utf-8"))
    assert len(data) == 1
    assert data[0]["item"]["nombre"] == "Casa"
    assert data[0]["item"]["precio"] == 100000


def test_record_save_csv(tmp_path):
    """record-save should save record as CSV file."""
    output = tmp_path / "test.csv"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string"  required="true"/>
                <record-set-field name="precio" type="integer" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
                <record-add-field name="precio">100000</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="csv" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-csv")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    content = output.read_text(encoding="utf-8")
    assert "item.nombre" in content
    assert "Casa" in content


def test_record_save_ndjson(tmp_path):
    """record-save should save record as NDJSON file."""
    output = tmp_path / "test.ndjson"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Depto</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="ndjson" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-ndjson")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    lines = output.read_text(encoding="utf-8").strip().split("\n")
    assert len(lines) == 2


def test_record_save_xml(tmp_path):
    """record-save should save record as XML."""
    output = tmp_path / "test.xml"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="xml" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-xml")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    text = output.read_text(encoding="utf-8")
    assert "Records" in text
    assert "Casa" in text


def test_record_save_xml_root_attrs_optional(tmp_path):
    """record-save xml: root attrs from record-create; per-record attrs on record-save."""
    output = tmp_path / "out.xml"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <xml-root-attr name="client">acme</xml-root-attr>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="xml" path="{output.as_posix()}"
                     xml-include-root-workflow="false"
                     xml-include-record-workflow="false"
                     xml-include-record-key="false">
            <xml-record-attr name="source">scraping</xml-record-attr>
        </record-save>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-attrs")

    assert session.status == SessionStatus.COMPLETED
    text = output.read_text(encoding="utf-8")
    assert 'total_records="1"' in text
    assert 'client="acme"' in text
    assert 'source="scraping"' in text
    assert 'workflow="' not in text
    assert "recordKey" not in text


def test_record_save_xml_built_in_root_attrs(tmp_path):
    """record-create xml-root-system enables session_id, francis_suite_version, exported_at on save."""
    output = tmp_path / "sys.xml"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <xml-root-system name="session_id"/>
            <xml-root-system name="francis_suite_version"/>
            <xml-root-system name="exported_at"/>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">X</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="xml" path="{output.as_posix()}"
                     xml-include-root-workflow="false"
                     xml-include-root-total-records="false"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-sys-attrs")

    assert session.status == SessionStatus.COMPLETED
    text = output.read_text(encoding="utf-8")
    assert "session_id=" in text
    assert "francis_suite_version=" in text
    assert "exported_at=" in text
    assert "total_records=" not in text


def test_record_save_json_export_from_record_create(tmp_path):
    """record-create record-export-* is embedded as _export on json (and other formats)."""
    output = tmp_path / "exp.json"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-export-attr name="client">acme</record-export-attr>
            <record-export-system name="francis_suite_version"/>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">X</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="json" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-json-export-children")

    assert session.status == SessionStatus.COMPLETED
    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["_export"]["client"] == "acme"
    assert "francis_suite_version" in data["_export"]
    assert len(data["data"]) == 1


def test_record_xml_record_attr_from_field(tmp_path):
    """record-xml-record-attr from-field: per-row value on <record>, independent of <Records> root."""
    output = tmp_path / "out.xml"

    xml = f"""
    <francis-workflow>
        <record-create name="r">
            <record-xml-record-attr name="id" from-field="item.record_key" required="true"/>
            <record-set-group name="item" required="true">
                <record-set-field name="record_key" type="string" required="true"/>
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="record_key">book-1</record-add-field>
                <record-add-field name="nombre">A</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="r" format="xml" path="{output.as_posix()}"
                     xml-include-root-workflow="false"
                     xml-include-record-workflow="false"
                     xml-include-record-key="false"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-record-attr")

    assert session.status == SessionStatus.COMPLETED
    text = output.read_text(encoding="utf-8")
    assert 'id="book-1"' in text


def test_record_xml_only_root_not_in_json(tmp_path):
    """record-xml-root-attr applies only to XML; JSON export does not include that key."""
    out_xml = tmp_path / "a.xml"
    out_json = tmp_path / "a.json"

    xml = f"""
    <francis-workflow>
        <record-create name="r">
            <record-xml-root-attr name="label">root-xml-only</record-xml-root-attr>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="nombre">X</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="r" format="xml" path="{out_xml.as_posix()}"
                     xml-include-root-workflow="false"
                     xml-include-root-total-records="false"/>
        <record-save from="r" format="json" path="{out_json.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-only-root")

    assert session.status == SessionStatus.COMPLETED
    assert 'label="root-xml-only"' in out_xml.read_text(encoding="utf-8")
    j = json.loads(out_json.read_text(encoding="utf-8"))
    assert "label" not in json.dumps(j)


def test_record_xml_required_record_attr_empty_fails(tmp_path):
    """required record-xml-record-attr with empty from-field value fails at save."""
    output = tmp_path / "x.xml"

    xml = f"""
    <francis-workflow>
        <record-create name="r">
            <record-xml-record-attr name="id" from-field="item.tag" required="true"/>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
                <record-set-field name="tag" type="string" required="false"/>
            </record-set-group>
        </record-create>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="nombre">A</record-add-field>
                <record-add-field name="tag"></record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="r" format="xml" path="{output.as_posix()}"
                     xml-include-root-workflow="false"
                     xml-include-record-workflow="false"
                     xml-include-record-key="false"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-xml-req-empty")

    assert session.status == SessionStatus.FAILED


def test_record_save_html(tmp_path):
    """record-save should save record as HTML table."""
    output = tmp_path / "test.html"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="html" path="{output.as_posix()}" html-title="Test"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-html")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    text = output.read_text(encoding="utf-8")
    assert "<table>" in text
    assert "Casa" in text
    assert "Test" in text


def test_record_save_txt(tmp_path):
    """record-save should save record as tab-separated text."""
    output = tmp_path / "test.txt"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="txt" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-txt")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()
    lines = output.read_text(encoding="utf-8").strip().split("\n")
    assert "item.nombre" in lines[0]
    assert "Casa" in lines[1]


def test_record_save_excel(tmp_path):
    """record-save should save record as xlsx."""
    output = tmp_path / "test.xlsx"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="excel" path="{output.as_posix()}" sheet-name="Items"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-excel")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()

    from openpyxl import load_workbook

    wb = load_workbook(output)
    assert "Items" in wb.sheetnames
    ws = wb["Items"]
    assert ws.cell(row=1, column=1).value == "item.nombre"
    assert ws.cell(row=2, column=1).value == "Casa"


def test_record_save_parquet(tmp_path):
    """record-save should save record as parquet."""
    output = tmp_path / "test.parquet"

    xml = f"""
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="testRecords" format="parquet" path="{output.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-save-parquet")

    assert session.status == SessionStatus.COMPLETED
    assert output.exists()

    import pyarrow.parquet as pq

    table = pq.read_table(output)
    assert table.num_rows == 1
    assert table.column("item.nombre")[0].as_py() == "Casa"


def test_record_add_without_create_fails():
    """record-add should fail if record was not created first."""
    xml = """
    <francis-workflow>
        <record-add to="noExiste">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-add-no-create")

    assert session.status == SessionStatus.FAILED


def test_record_add_multiple_rows():
    """record-add should accumulate multiple rows."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Depto</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Local</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()

    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-multiple-rows")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.count == 3
    assert record.last_row["item"]["nombre"] == "Local"


def test_record_key_duplicate_skipped():
    """record-add with same key-field values should append once when record-key is set."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-key>
                <key-field name="nombre"/>
            </record-key>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
                <record-set-field name="precio" type="integer" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
                <record-add-field name="precio">100</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="nombre">Casa</record-add-field>
                <record-add-field name="precio">200</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-key-dup")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert isinstance(record, FRecord)
    assert record.count == 1
    assert record.last_row["item"]["precio"] == 100


def test_record_key_two_fields():
    """Two different keys should both be stored."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-key>
                <key-field name="a"/>
                <key-field name="b"/>
            </record-key>
            <record-set-group name="item" required="true">
                <record-set-field name="a" type="string" required="true"/>
                <record-set-field name="b" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="a">1</record-add-field>
                <record-add-field name="b">2</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="a">1</record-add-field>
                <record-add-field name="b">3</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-key-2")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.count == 2


def test_record_key_qualified_name():
    """key-field name='group.field' when bare name would be ambiguous — use one group."""
    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-key>
                <key-field name="item.id"/>
            </record-key>
            <record-set-group name="item" required="true">
                <record-set-field name="id" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="testRecords">
            <record-add-group name="item">
                <record-add-field name="id">x</record-add-field>
            </record-add-group>
        </record-add>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-key-qualified")

    assert session.status == SessionStatus.COMPLETED
    from francis_suite.core.records import FRecord
    record = session.context.get_shared_box("testRecords")
    assert record.count == 1


def test_auto_private_record_metadata_without_hand(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    monkeypatch.setenv("FRANCIS_AUTO_RECORD_METADATA", "1")

    xml = """
    <francis-workflow>
        <record-create name="testRecords">
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-auto-private-meta")

    assert session.status == SessionStatus.COMPLETED
    meta_path = tmp_path / "sessions" / session.id / "testRecords_private_metadata.json"
    assert meta_path.exists()
    import json
    data = json.loads(meta_path.read_text(encoding="utf-8"))
    assert data.get("session_id") == session.id
    assert data.get("status") == "completed"


def test_record_journal_one_line_per_add(tmp_path):
    """record-journal appends one NDJSON line per successful record-add."""
    journal = tmp_path / "rows.ndjson"
    out = tmp_path / "out.json"

    xml = f"""
    <francis-workflow>
        <record-create name="r">
            <record-journal path="{journal.as_posix()}"/>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="nombre">A</record-add-field>
            </record-add-group>
        </record-add>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="nombre">B</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="r" format="json" path="{out.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-record-journal")

    assert session.status == SessionStatus.COMPLETED
    lines = journal.read_text(encoding="utf-8").strip().split("\n")
    assert len(lines) == 4
    assert json.loads(lines[0])["_type"] == "journal_header"
    assert json.loads(lines[1])["_type"] == "record"
    assert json.loads(lines[1])["data"]["item.nombre"] == "A"
    assert json.loads(lines[2])["data"]["item.nombre"] == "B"
    assert json.loads(lines[3])["_type"] == "process"
    assert json.loads(lines[3])["status"] == "completed"


def test_record_journal_process_line_on_failed_session(tmp_path):
    """After a failed workflow, journal ends with _type process and status failed."""
    journal = tmp_path / "j.ndjson"

    xml = f"""
    <francis-workflow>
        <record-create name="r">
            <record-journal path="{journal.as_posix()}"/>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="nombre">only-row</record-add-field>
            </record-add-group>
        </record-add>
        <tag-que-no-existe/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-journal-fail")

    assert session.status == SessionStatus.FAILED
    lines = journal.read_text(encoding="utf-8").strip().split("\n")
    assert len(lines) >= 3
    assert json.loads(lines[-1])["_type"] == "process"
    assert json.loads(lines[-1])["status"] == "failed"
    assert json.loads(lines[-1])["rows_committed"] == 1


def test_record_export_system_status_hide_when_completed(tmp_path):
    """show-attribute=false on status_process omits key when session completed."""
    out = tmp_path / "out.json"

    xml = f"""
    <francis-workflow>
        <record-create name="r">
            <record-export-system name="status_process" show-attribute="false"/>
            <record-set-group name="item" required="true">
                <record-set-field name="nombre" type="string" required="true"/>
            </record-set-group>
        </record-create>
        <record-add to="r">
            <record-add-group name="item">
                <record-add-field name="nombre">X</record-add-field>
            </record-add-group>
        </record-add>
        <record-save from="r" format="json" path="{out.as_posix()}"/>
    </francis-workflow>
    """

    parser = FParser()
    runtime = FRuntime()
    root = parser.parse_string(xml)
    session = runtime.run(root, workflow_name="test-status-hide")

    assert session.status == SessionStatus.COMPLETED
    data = json.loads(out.read_text(encoding="utf-8"))
    assert "status_process" not in data["_export"]
